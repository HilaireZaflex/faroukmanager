#!/usr/bin/env python3
"""
Script d'import des vrais PDV depuis le fichier Excel BASE FaroukManager.xlsx
Importe la feuille BASE REELLE et enrichit avec NOUVELLE ATTRIBUTION SIM
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from datetime import datetime
from app.core.database import SessionLocal, Base, engine
import app.models  # noqa
from app.models.pdv import PDV, PDVType, PDVStatut, PDVMedaille

EXCEL_PATH = "/Users/nms/Downloads/BASE FaroukManager.xlsx"

def clean_str(val):
    """Nettoie une valeur string"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'None', 'nan', 'AU BUREAU', '=SUBTOTAL(103,Tableau13[TYPE])',
             '=SUBTOTAL(103,Tableau13[ZONE])', '=SUBTOTAL(103,Tableau13[GESTIONNAIRES])',
             '=SUBTOTAL(103,Tableau13[TELECONSEILLIERE])', '=SUBTOTAL(103,Tableau13[SINGLE WALET])',
             '=SUBTOTAL(103,Tableau13[DEVELOPPEUR])'):
        return None
    return s

def get_type_pdv(val):
    """Détermine le type PDV"""
    v = clean_str(val)
    if v is None:
        return PDVType.RS
    v = v.upper().strip()
    mapping = {
        'KIOSQUE': PDVType.KIOSQUE,
        'KIOSQUE ': PDVType.KIOSQUE,
        'KIOSQUE INDEPENDANT': PDVType.KIOSQUE_INDEPENDANT,
        'RNS': PDVType.RNS,
        'RS': PDVType.RS,
        'RSF': PDVType.RSF,
        'NEANT': PDVType.NEANT,
        'X': PDVType.X,
    }
    return mapping.get(v, PDVType.RS)

def get_statut(single_wallet, commentaire):
    """Détermine le statut du PDV"""
    sw = clean_str(single_wallet)
    comm = clean_str(commentaire)

    # Si commentaire contient SIM RECUPEREE → RECUPERATION
    if comm and 'SIM RECUPEREE' in comm.upper():
        return PDVStatut.RECUPERATION

    if sw is None or sw.upper() == 'NEANT':
        return PDVStatut.INACTIF
    if sw.upper() == 'OPERAT':
        return PDVStatut.ACTIF
    if 'NOUVELLE CREAT' in sw.upper():
        return PDVStatut.ACTIF
    return PDVStatut.ACTIF

def get_single_wallet_flags(single_wallet):
    """Retourne les flags liés au single wallet"""
    sw = clean_str(single_wallet)
    nouvelle_creation = False
    if sw and 'NOUVELLE CREAT' in sw.upper():
        nouvelle_creation = True
    return nouvelle_creation

def main():
    print("📂 Chargement du fichier Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Charger NOUVELLE ATTRIBUTION SIM ---
    ws_attrib = wb['NOUVELLE ATTRIBUTION SIM']
    nouvelles_attribs = {}
    for row in ws_attrib.iter_rows(min_row=2, values_only=True):
        pdv_num = row[0]
        if pdv_num is None:
            continue
        nouvelles_attribs[str(pdv_num).strip()] = {
            'type': row[1],
            'date_attribution': row[2],
            'dev_recu': row[3],
            'dev_active': row[4],
            'date_activation_attrib': row[5],
            'date_retour_fiche': row[6],
            'conditions': row[7],
            'intervalle': row[8],
        }
    print(f"✅ {len(nouvelles_attribs)} nouvelles attributions chargées")

    # --- Charger BASE REELLE ---
    ws = wb['BASE REELLE']
    print("🔄 Import des PDV en cours...")

    # Recréer les tables
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    print("🗑️  Base de données réinitialisée")

    db = SessionLocal()
    imported = 0
    skipped = 0
    au_bureau = 0

    try:
        # Recréer l'admin
        from app.models.user import User, UserRole
        from app.core.security import get_password_hash
        admin = User(
            email="admin@faroukmanager.com",
            nom="Administrateur",
            hashed_password=get_password_hash("Admin2026!"),
            role=UserRole.ADMIN,
            is_active=True
        )
        db.add(admin)
        db.commit()
        print("✅ Admin recréé")

        seen_pdvs = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            localite    = row[0]
            adresse     = row[1]
            nom_complet = row[2]
            num_perso   = row[3]
            pdv_num     = row[4]
            type_pdv    = row[5]
            single_wall = row[6]
            date_activ  = row[7]
            superviseur = row[8]
            gestionnaire= row[9]
            zone        = row[10]
            sous_zone   = row[11]
            telecons    = row[12]
            developpeur = row[13]
            commentaire = row[14]
            date_maj    = row[15]

            # Ignorer les lignes vides ou formules
            if pdv_num is None:
                skipped += 1
                continue

            pdv_str = str(pdv_num).strip()

            # Ignorer les formules subtotal
            if pdv_str.startswith('='):
                skipped += 1
                continue

            # Compter AU BUREAU mais les importer quand même
            if clean_str(localite) is None and clean_str(nom_complet) is None:
                au_bureau += 1

            # Éviter les doublons
            if pdv_str in seen_pdvs:
                skipped += 1
                continue
            seen_pdvs.add(pdv_str)

            # Récupérer les infos de nouvelle attribution si disponible
            attrib = nouvelles_attribs.get(pdv_str, {})

            # Date d'activation : priorité à la feuille BASE REELLE
            date_activation = None
            if isinstance(date_activ, datetime):
                date_activation = date_activ
            elif attrib.get('date_activation_attrib') and isinstance(attrib['date_activation_attrib'], datetime):
                date_activation = attrib['date_activation_attrib']

            # Ignorer dates aberrantes (ex: 1990)
            if date_activation and date_activation.year < 2000:
                date_activation = None

            # Nom du PDV
            nom = clean_str(nom_complet)
            if nom is None:
                nom = f"PDV {pdv_str}"

            # Numéro personnel
            num_p = clean_str(str(num_perso)) if num_perso else None

            # Statut
            statut = get_statut(single_wall, commentaire)

            # SIM au bureau
            sim_au_bureau = (clean_str(localite) is None and clean_str(nom_complet) is None)

            # Nouvelle création
            nouvelle_creation = get_single_wallet_flags(single_wall)

            # Notes : combiner commentaire + infos attribution
            notes_parts = []
            if clean_str(commentaire):
                notes_parts.append(f"Commentaire: {clean_str(commentaire)}")
            if attrib.get('conditions') and clean_str(str(attrib['conditions'])):
                notes_parts.append(f"Conditions activation: {attrib['conditions']}")
            if attrib.get('intervalle') and clean_str(str(attrib['intervalle'])):
                notes_parts.append(f"Délai retour fiche: {attrib['intervalle']}")
            if date_maj:
                notes_parts.append(f"Dernière MAJ: {date_maj}")
            notes = " | ".join(notes_parts) if notes_parts else None

            # Superviseur : nettoyer (certains ont "AU BUREAU ZONE X")
            sup = clean_str(superviseur)
            if sup and sup.startswith('AU BUREAU'):
                sup = None

            pdv = PDV(
                numero_pdv=pdv_str,
                nom=nom,
                nom_gerant=nom,
                numero_personnel=num_p,
                telephone=num_p,
                type_pdv=get_type_pdv(type_pdv),
                statut=statut,
                medaille=PDVMedaille.AUCUNE,
                zone=clean_str(str(zone)) if zone else None,
                sous_zone=clean_str(str(sous_zone)) if sous_zone else None,
                quartier=clean_str(str(localite)) if localite else None,
                commune=clean_str(str(localite)) if localite else None,
                adresse=clean_str(str(adresse)) if adresse else None,
                superviseur=sup,
                gestionnaire=clean_str(str(gestionnaire)) if gestionnaire else None,
                teleconseillere=clean_str(str(telecons)) if telecons else None,
                developpeur=clean_str(str(developpeur)) if developpeur else None,
                date_activation=date_activation,
                sim_au_bureau=sim_au_bureau,
                nouvelle_creation=nouvelle_creation,
                health_score=50.0,
                score_risque=0.0,
                notes=notes,
            )
            db.add(pdv)
            imported += 1

            if imported % 100 == 0:
                db.commit()
                print(f"  ... {imported} PDV importés")

        db.commit()
        print()
        print(f"✅ Import terminé !")
        print(f"   📊 PDV importés    : {imported}")
        print(f"   📋 SIM au bureau   : {au_bureau}")
        print(f"   ⏭️  Lignes ignorées : {skipped}")

        # Vérification rapide
        total = db.query(PDV).count()
        actifs = db.query(PDV).filter(PDV.statut == PDVStatut.ACTIF).count()
        inactifs = db.query(PDV).filter(PDV.statut == PDVStatut.INACTIF).count()
        recup = db.query(PDV).filter(PDV.statut == PDVStatut.RECUPERATION).count()
        print()
        print(f"📈 Statistiques DB:")
        print(f"   Total PDV : {total}")
        print(f"   Actifs    : {actifs}")
        print(f"   Inactifs  : {inactifs}")
        print(f"   En récup. : {recup}")

    except Exception as e:
        db.rollback()
        print(f"❌ Erreur : {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()

if __name__ == "__main__":
    main()
