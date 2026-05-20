#!/usr/bin/env python3
"""
Script d'import depuis le fichier EXPORT Orange (EXPORT_XXXXXXXXX.xlsx)
==========================================================================
Structure du fichier Orange :
  A: Numero revendeur    → numéro PDV
  B: Grade               → type PDV
  C: Numero parent       → superviseur
  D: Numero pdg          → notre numéro réseau
  E: Nom pdg             → nom réseau
  F: Service             → CASHIN / CASHOUT / TXNCORRECT / B2BCASHOUT
  G: Nombre transaction  → nb opérations pour ce service/jour
  H: Montant transaction → volume financier
  I: Transaction CA      → CA Orange (base CASHOUT uniquement)
  J: Commission PDG      → commission que le réseau reçoit
  K: Commission revendeur→ commission que le PDV reçoit
  L: Tax amount          → taxes
  M: Prelevement         → prélèvements
  N: Date transaction    → date (format DD/MM/YYYY)

Agrégation :
  - Grouper par (Numero revendeur, mois/semaine)
  - nb_depots        = Σ(Nombre transaction) où Service = CASHIN
  - montant_depots   = Σ(Montant transaction) où Service = CASHIN
  - nb_retraits      = Σ(Nombre transaction) où Service = CASHOUT
  - montant_retraits = Σ(Montant transaction) où Service = CASHOUT
  - montant_transaction = montant_depots + montant_retraits
  - montant_ca       = Σ(Transaction CA) tous services
  - commission_pdg   = Σ(Commission PDG) tous services
  - commission_revendeur = Σ(Commission revendeur) tous services
  - ratio_ca_transaction = montant_ca / montant_transaction × 100

Usage :
  python3 import_export_orange.py --file /path/to/EXPORT_XXXX.xlsx --mode mensuel
  python3 import_export_orange.py --file /path/to/EXPORT_XXXX.xlsx --mode hebdo
  python3 import_export_orange.py --file /path/to/EXPORT_XXXX.xlsx --mode mensuel --annee 2026 --mois 4
"""
import sys
import os
import argparse
from datetime import datetime
from collections import defaultdict

# Positionner dans le répertoire backend
backend_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'backend')
os.chdir(backend_dir)
sys.path.insert(0, backend_dir)

import openpyxl
from app.core.database import SessionLocal
from app.models.pdv import PDV
from app.models.performance import MonthlyPerformance, WeeklyPerformance


def parse_float(val):
    """Convertit une valeur en float, retourne 0.0 si None/vide."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(' ', '').replace(',', '.'))
    except (ValueError, TypeError):
        return 0.0


def parse_int(val):
    """Convertit une valeur en int, retourne 0 si None/vide."""
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(' ', '').replace(',', '.')))
    except (ValueError, TypeError):
        return 0


def parse_date(val):
    """Parse une date au format DD/MM/YYYY ou YYYY-MM-DD."""
    if val is None:
        return None
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def get_semaine(d):
    """Retourne le numéro de semaine ISO d'une date."""
    return d.isocalendar()[1]


def load_excel(filepath):
    """Charge le fichier EXPORT Orange et retourne les lignes."""
    print(f"📂 Chargement : {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    # Lecture des en-têtes (ligne 1)
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell or '').strip().lower())

    print(f"   Colonnes détectées : {headers}")

    # Mapping colonnes → index
    col_map = {
        'numero_revendeur': None, 'grade': None, 'numero_parent': None,
        'numero_pdg': None, 'nom_pdg': None, 'service': None,
        'nombre_transaction': None, 'montant_transaction': None,
        'transaction_ca': None, 'commission_pdg': None, 'commission_revendeur': None,
        'tax_amount': None, 'prelevement': None, 'date_transaction': None,
    }

    # Correspondances flexibles
    header_aliases = {
        'numero revendeur':     'numero_revendeur',
        'grade':                'grade',
        'numero parent':        'numero_parent',
        'numero pdg':           'numero_pdg',
        'nom pdg':              'nom_pdg',
        'service':              'service',
        'nombre transction':    'nombre_transaction',
        'nombre transaction':   'nombre_transaction',
        'montant transaction':  'montant_transaction',
        'transaction ca':       'transaction_ca',
        'commission pdg':       'commission_pdg',
        'commission revendeur': 'commission_revendeur',
        'tax amount':           'tax_amount',
        'prélevement':          'prelevement',
        'prelevement':          'prelevement',
        'date transaction':     'date_transaction',
    }

    for i, h in enumerate(headers):
        key = header_aliases.get(h)
        if key:
            col_map[key] = i

    # Vérification colonnes obligatoires
    required = ['numero_revendeur', 'service', 'nombre_transaction',
                'montant_transaction', 'date_transaction']
    missing = [r for r in required if col_map[r] is None]
    if missing:
        print(f"❌ Colonnes obligatoires manquantes : {missing}")
        print(f"   Colonnes trouvées : {headers}")
        sys.exit(1)

    print(f"   ✅ Mapping colonnes OK")
    return ws, col_map, headers


def aggregate_rows(ws, col_map, mode, annee_filter=None, mois_filter=None, semaine_filter=None):
    """
    Agrège les lignes du fichier par (PDV, période).
    Retourne un dict: key → données agrégées.

    key = (numero_pdv, annee, mois)    si mode='mensuel'
    key = (numero_pdv, annee, semaine) si mode='hebdo'
    """
    print(f"🔄 Agrégation des données (mode={mode})...")

    # Structure d'agrégation
    data = defaultdict(lambda: {
        'nb_depots': 0, 'montant_depots': 0.0,
        'nb_retraits': 0, 'montant_retraits': 0.0,
        'nb_b2bcashout': 0, 'montant_b2bcashout': 0.0,
        'nb_txncorrect': 0, 'montant_txncorrect': 0.0,
        'montant_transaction': 0.0,
        'montant_ca': 0.0,
        'commission_pdg': 0.0,
        'commission_revendeur': 0.0,
        'grade': None, 'dates': set(),
    })

    skipped_date = 0
    skipped_pdv = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col_map['numero_revendeur']] is None:
            continue

        total_rows += 1
        numero = str(row[col_map['numero_revendeur']]).strip()

        # Parse date
        date_val = row[col_map['date_transaction']]
        d = parse_date(date_val)
        if d is None:
            skipped_date += 1
            continue

        annee = d.year
        mois  = d.month
        semaine = get_semaine(d)

        # Filtres optionnels
        if annee_filter and annee != annee_filter:
            continue
        if mois_filter and mois != mois_filter:
            continue
        if semaine_filter and semaine != semaine_filter:
            continue

        # Clé d'agrégation
        if mode == 'mensuel':
            key = (numero, annee, mois)
        else:
            key = (numero, annee, semaine)

        service = str(row[col_map['service']] or '').strip().upper()
        nb_trans    = parse_int(row[col_map['nombre_transaction']])
        mt_trans    = parse_float(row[col_map['montant_transaction']])
        trans_ca    = parse_float(row[col_map['transaction_ca']])    if col_map['transaction_ca'] is not None else 0.0
        comm_pdg    = parse_float(row[col_map['commission_pdg']])    if col_map['commission_pdg'] is not None else 0.0
        comm_rev    = parse_float(row[col_map['commission_revendeur']]) if col_map['commission_revendeur'] is not None else 0.0
        grade       = str(row[col_map['grade']] or '').strip() if col_map['grade'] is not None else None

        rec = data[key]
        if grade and not rec['grade']:
            rec['grade'] = grade

        rec['dates'].add(d.date())
        rec['montant_ca']          += trans_ca
        rec['commission_pdg']      += comm_pdg
        rec['commission_revendeur'] += comm_rev

        if service == 'CASHIN':
            rec['nb_depots']        += nb_trans
            rec['montant_depots']   += mt_trans
        elif service == 'CASHOUT':
            rec['nb_retraits']      += nb_trans
            rec['montant_retraits'] += mt_trans
        elif service == 'B2BCASHOUT':
            rec['nb_b2bcashout']    += nb_trans
            rec['montant_b2bcashout'] += mt_trans
        elif service == 'TXNCORRECT':
            rec['nb_txncorrect']    += nb_trans
            rec['montant_txncorrect'] += mt_trans

    # Calculer montant_transaction et ratio
    for key, rec in data.items():
        rec['montant_transaction'] = rec['montant_depots'] + rec['montant_retraits']
        mt = rec['montant_transaction']
        rec['ratio_ca_transaction'] = round(
            (rec['montant_ca'] / mt * 100) if mt > 0 else 0.0, 4
        )
        rec['nb_operations'] = rec['nb_depots'] + rec['nb_retraits']
        rec['est_actif'] = rec['nb_operations'] > 0

    print(f"   Total lignes lues    : {total_rows:,}")
    print(f"   PDV × périodes       : {len(data):,}")
    print(f"   Dates invalides      : {skipped_date}")
    return data


def import_mensuel(filepath, annee_filter=None, mois_filter=None):
    """Import mensuel depuis le fichier EXPORT Orange."""
    ws, col_map, _ = load_excel(filepath)
    data = aggregate_rows(ws, col_map, 'mensuel', annee_filter, mois_filter)

    db = SessionLocal()
    try:
        # Charger PDV map
        pdvs = db.query(PDV).all()
        pdv_map = {str(p.numero_pdv).strip(): p.id for p in pdvs}
        print(f"   {len(pdv_map)} PDV chargés en base")

        created = updated = skipped = 0
        not_found = set()

        for (numero, annee, mois), rec in data.items():
            pdv_id = pdv_map.get(numero)
            if pdv_id is None:
                not_found.add(numero)
                skipped += 1
                continue

            mt = rec['montant_transaction']
            ca = rec['montant_ca']

            # Upsert
            existing = db.query(MonthlyPerformance).filter_by(
                pdv_id=pdv_id, annee=annee, mois=mois
            ).first()

            if existing:
                existing.nb_depots          = rec['nb_depots']
                existing.montant_depots     = rec['montant_depots']
                existing.nb_retraits        = rec['nb_retraits']
                existing.montant_retraits   = rec['montant_retraits']
                existing.nb_operations      = rec['nb_operations']
                existing.montant_transaction = mt
                existing.montant_ca         = ca
                existing.commission_pdg     = rec['commission_pdg']
                existing.commission_revendeur = rec['commission_revendeur']
                existing.ratio_ca_transaction = rec['ratio_ca_transaction']
                existing.est_actif          = rec['est_actif']
                # Synchroniser ca (compatibilité) = montant_transaction
                existing.ca                 = mt
                updated += 1
            else:
                db.add(MonthlyPerformance(
                    pdv_id=pdv_id, annee=annee, mois=mois,
                    nb_depots=rec['nb_depots'],
                    montant_depots=rec['montant_depots'],
                    nb_retraits=rec['nb_retraits'],
                    montant_retraits=rec['montant_retraits'],
                    nb_operations=rec['nb_operations'],
                    montant_transaction=mt,
                    montant_ca=ca,
                    commission_pdg=rec['commission_pdg'],
                    commission_revendeur=rec['commission_revendeur'],
                    ratio_ca_transaction=rec['ratio_ca_transaction'],
                    est_actif=rec['est_actif'],
                    ca=mt,  # compatibilité
                ))
                created += 1

            if (created + updated) % 200 == 0:
                db.commit()
                print(f"   ... {created + updated} enregistrements traités")

        db.commit()

        # Calculer les variations mois/mois
        print("📊 Calcul des variations mensuelles...")
        perfs = db.query(MonthlyPerformance).order_by(
            MonthlyPerformance.pdv_id,
            MonthlyPerformance.annee,
            MonthlyPerformance.mois
        ).all()
        prev = {}
        for p in perfs:
            k = p.pdv_id
            mt_cur = p.montant_transaction or p.ca or 0
            if k in prev and prev[k] > 0:
                p.ca_mois_precedent = prev[k]
                p.taux_variation = round(((mt_cur - prev[k]) / prev[k]) * 100, 2)
            prev[k] = mt_cur
        db.commit()
        print("   ✅ Variations calculées")

        # Résumé
        print()
        print(f"✅ Import mensuel terminé !")
        print(f"   Créés   : {created}")
        print(f"   Mis à jour : {updated}")
        print(f"   Ignorés (PDV non trouvés) : {skipped} ({len(not_found)} uniques)")
        if not_found:
            print(f"   Premiers PDV non trouvés : {list(not_found)[:10]}")

        # Stats par période
        from sqlalchemy import func
        stats = db.query(
            MonthlyPerformance.annee,
            MonthlyPerformance.mois,
            func.count(MonthlyPerformance.id).label('nb'),
            func.sum(MonthlyPerformance.montant_transaction).label('mt_total'),
            func.sum(MonthlyPerformance.montant_ca).label('ca_total'),
            func.sum(MonthlyPerformance.commission_pdg).label('comm_pdg'),
            func.sum(MonthlyPerformance.commission_revendeur).label('comm_rev'),
        ).group_by(MonthlyPerformance.annee, MonthlyPerformance.mois)\
         .order_by(MonthlyPerformance.annee, MonthlyPerformance.mois).all()

        MOIS = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
                7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}
        print()
        print("📈 Résumé par période :")
        for s in stats:
            print(f"   {MOIS[s.mois]} {s.annee}: {s.nb} PDV | "
                  f"Transaction={s.mt_total:>15,.0f} | "
                  f"CA={s.ca_total:>12,.0f} | "
                  f"CommPDG={s.comm_pdg:>10,.2f} | "
                  f"CommRev={s.comm_rev:>10,.2f}")

    except Exception as e:
        db.rollback()
        print(f"❌ Erreur : {e}")
        import traceback; traceback.print_exc()
        raise
    finally:
        db.close()


def import_hebdo(filepath, annee_filter=None, semaine_filter=None):
    """Import hebdomadaire depuis le fichier EXPORT Orange."""
    ws, col_map, _ = load_excel(filepath)
    data = aggregate_rows(ws, col_map, 'hebdo', annee_filter, semaine_filter=semaine_filter)

    db = SessionLocal()
    try:
        pdvs = db.query(PDV).all()
        pdv_map = {str(p.numero_pdv).strip(): p.id for p in pdvs}
        print(f"   {len(pdv_map)} PDV chargés en base")

        created = updated = skipped = 0
        not_found = set()

        for (numero, annee, semaine), rec in data.items():
            pdv_id = pdv_map.get(numero)
            if pdv_id is None:
                not_found.add(numero)
                skipped += 1
                continue

            mt = rec['montant_transaction']

            existing = db.query(WeeklyPerformance).filter_by(
                pdv_id=pdv_id, annee=annee, semaine=semaine
            ).first()

            if existing:
                existing.nb_depots           = rec['nb_depots']
                existing.montant_depots      = rec['montant_depots']
                existing.nb_retraits         = rec['nb_retraits']
                existing.montant_retraits    = rec['montant_retraits']
                existing.nb_operations       = rec['nb_operations']
                existing.montant_transaction = mt
                existing.montant_ca          = rec['montant_ca']
                existing.commission_pdg      = rec['commission_pdg']
                existing.commission_revendeur = rec['commission_revendeur']
                existing.ratio_ca_transaction = rec['ratio_ca_transaction']
                existing.est_actif           = rec['est_actif']
                existing.ca                  = mt
                updated += 1
            else:
                db.add(WeeklyPerformance(
                    pdv_id=pdv_id, annee=annee, semaine=semaine,
                    nb_depots=rec['nb_depots'],
                    montant_depots=rec['montant_depots'],
                    nb_retraits=rec['nb_retraits'],
                    montant_retraits=rec['montant_retraits'],
                    nb_operations=rec['nb_operations'],
                    montant_transaction=mt,
                    montant_ca=rec['montant_ca'],
                    commission_pdg=rec['commission_pdg'],
                    commission_revendeur=rec['commission_revendeur'],
                    ratio_ca_transaction=rec['ratio_ca_transaction'],
                    est_actif=rec['est_actif'],
                    ca=mt,
                ))
                created += 1

            if (created + updated) % 200 == 0:
                db.commit()

        db.commit()

        # Variations semaine/semaine
        print("📊 Calcul des variations hebdomadaires...")
        perfs = db.query(WeeklyPerformance).order_by(
            WeeklyPerformance.pdv_id,
            WeeklyPerformance.annee,
            WeeklyPerformance.semaine
        ).all()
        prev = {}
        for p in perfs:
            k = p.pdv_id
            mt_cur = p.montant_transaction or p.ca or 0
            if k in prev and prev[k] > 0:
                p.ca_semaine_precedente = prev[k]
                p.taux_variation = round(((mt_cur - prev[k]) / prev[k]) * 100, 2)
            prev[k] = mt_cur
        db.commit()

        print()
        print(f"✅ Import hebdomadaire terminé !")
        print(f"   Créés   : {created}")
        print(f"   Mis à jour : {updated}")
        print(f"   Ignorés : {skipped} ({len(not_found)} PDV non trouvés)")

    except Exception as e:
        db.rollback()
        print(f"❌ Erreur : {e}")
        import traceback; traceback.print_exc()
        raise
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(
        description='Import fichier EXPORT Orange → FaroukManager'
    )
    parser.add_argument('--file', required=True, help='Chemin vers le fichier EXPORT_XXXX.xlsx')
    parser.add_argument('--mode', choices=['mensuel', 'hebdo'], default='mensuel',
                        help='Mode d\'import : mensuel ou hebdo (défaut: mensuel)')
    parser.add_argument('--annee', type=int, default=None, help='Filtrer sur une année')
    parser.add_argument('--mois', type=int, default=None, help='Filtrer sur un mois (1-12)')
    parser.add_argument('--semaine', type=int, default=None, help='Filtrer sur une semaine (1-52)')
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"❌ Fichier introuvable : {args.file}")
        sys.exit(1)

    print(f"🚀 FaroukManager — Import EXPORT Orange")
    print(f"   Fichier : {args.file}")
    print(f"   Mode    : {args.mode}")
    if args.annee:  print(f"   Année   : {args.annee}")
    if args.mois:   print(f"   Mois    : {args.mois}")
    if args.semaine: print(f"   Semaine : {args.semaine}")
    print()

    if args.mode == 'mensuel':
        import_mensuel(args.file, args.annee, args.mois)
    else:
        import_hebdo(args.file, args.annee, args.semaine)


if __name__ == "__main__":
    main()
