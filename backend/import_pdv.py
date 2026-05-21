"""
Script d'importation des PDVs depuis BASE.xlsx
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from datetime import datetime
from app.core.database import SessionLocal, engine
from app.models.pdv import PDV, PDVType, PDVStatut, Base

# Créer les tables si elles n'existent pas
Base.metadata.create_all(bind=engine)

FICHIER = '/Users/nms/Downloads/IMPORTER/BASE.xlsx'

def normaliser_type(t):
    if not t:
        return PDVType.RS
    t = str(t).strip().upper()
    mapping = {
        'KIOSQUE': PDVType.KIOSQUE,
        'KIOSQUE INDEPENDANT': PDVType.KIOSQUE_INDEPENDANT,
        'KIOSQUE INDÉPENDANT': PDVType.KIOSQUE_INDEPENDANT,
        'RNS': PDVType.RNS,
        'RS': PDVType.RS,
        'RSF': PDVType.RSF,
        'NEANT': PDVType.NEANT,
        'NÉANT': PDVType.NEANT,
        'X': PDVType.X,
    }
    return mapping.get(t, PDVType.RS)

def normaliser_date(d):
    if not d:
        return None
    if isinstance(d, datetime):
        return d
    if isinstance(d, str):
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
            try:
                return datetime.strptime(d.strip(), fmt)
            except:
                pass
    return None

def est_au_bureau(val):
    return str(val).strip().upper() == 'AU BUREAU' if val else False

wb = openpyxl.load_workbook(FICHIER, read_only=True, data_only=True)

# ── Feuille 1 : données principales ─────────────────────────────────────────
ws1 = wb['LISTE DES INFOS SUR LES PDV']
rows1 = list(ws1.iter_rows(values_only=True))
headers1 = rows1[0]
print(f"📋 Feuille 1: {len(rows1)-1} PDVs à importer")

# ── Feuille NEW CREATION : MSISDN, IMSI, ETAT ───────────────────────────────
ws2 = wb['NEW CREATION']
rows2 = list(ws2.iter_rows(values_only=True))
new_creation_map = {}
for row in rows2[1:]:
    if row[0]:
        new_creation_map[str(row[0])] = {
            'msisdn': row[0],
            'imsi': row[1],
            'etat': row[2],
            'commentaire_nc': row[3],
        }
print(f"📋 NEW CREATION: {len(new_creation_map)} entrées")

# ── Feuille NOUVELLE ATTRIBUTION SIM ────────────────────────────────────────
ws3 = wb['NOUVELLE ATTRIBUTION SIM']
rows3 = list(ws3.iter_rows(values_only=True))
attribution_map = {}
for row in rows3[1:]:
    if row[0]:
        attribution_map[str(row[0])] = {
            'date_attribution': row[2],
            'dev_recu_sim': row[3],
            'dev_active_sim': row[4],
            'date_activation_pdv': row[5],
            'date_retour_fiche': row[6],
            'conditions_activation': row[7],
            'interval_activation': row[8],
        }
print(f"📋 NOUVELLE ATTRIBUTION SIM: {len(attribution_map)} entrées")

# ── Import dans la base ──────────────────────────────────────────────────────
db = SessionLocal()
inseres = 0
ignores = 0
erreurs = 0

for row in rows1[1:]:
    try:
        localite      = row[0]
        adresse       = row[1]
        nom           = row[2]
        num_personnel = row[3]
        num_pdv       = row[4]
        type_pdv      = row[5]
        single_wallet = row[6]
        date_activ    = row[7]
        superviseur   = row[8]
        gestionnaire  = row[9]
        zone          = row[10]
        sous_zone     = row[11]
        teleconseillere = row[12]
        developpeur   = row[13]
        commentaire   = row[14]
        date_maj      = row[15]

        if not num_pdv:
            ignores += 1
            continue

        num_pdv_str = str(int(num_pdv)) if isinstance(num_pdv, float) else str(num_pdv).strip()

        # Vérifier si déjà existant
        existing = db.query(PDV).filter(PDV.numero_pdv == num_pdv_str).first()
        if existing:
            ignores += 1
            continue

        # Statut selon contenu
        sim_bureau = est_au_bureau(localite)
        
        # Déterminer statut
        statut = PDVStatut.ACTIF
        single_wallet_str = str(single_wallet).strip().upper() if single_wallet else ''
        if 'RECUPER' in str(adresse).upper() or 'SIM RECUPER' in str(commentaire).upper():
            statut = PDVStatut.RECUPERATION
        elif single_wallet_str in ['PROBLEME DE DROIT', 'INACTIF', 'DESACTIVE']:
            statut = PDVStatut.INACTIF

        # Enrichir avec NEW CREATION si disponible
        nc = new_creation_map.get(num_pdv_str, {})
        attr = attribution_map.get(num_pdv_str, {})

        # Notes combinées
        notes_parts = []
        if commentaire and str(commentaire).strip() not in ['', 'None']:
            notes_parts.append(str(commentaire))
        if nc.get('commentaire_nc'):
            notes_parts.append(f"NC: {nc['commentaire_nc']}")
        if attr.get('conditions_activation'):
            notes_parts.append(f"Activation: {attr['conditions_activation']}")
        if attr.get('interval_activation'):
            notes_parts.append(f"Délai fiche: {attr['interval_activation']}")

        # Date d'activation
        date_activation = normaliser_date(date_activ)
        if not date_activation and attr.get('date_activation_pdv'):
            date_activation = normaliser_date(attr['date_activation_pdv'])

        pdv = PDV(
            numero_pdv      = num_pdv_str,
            nom             = str(nom).strip() if nom and str(nom).strip() != 'AU BUREAU' else num_pdv_str,
            numero_personnel= str(num_personnel).strip() if num_personnel and str(num_personnel).strip() != 'AU BUREAU' else None,
            type_pdv        = normaliser_type(type_pdv),
            statut          = statut,
            zone            = str(zone).strip() if zone and str(zone).strip() != 'AU BUREAU' else None,
            sous_zone       = str(sous_zone).strip() if sous_zone and str(sous_zone).strip() != 'AU BUREAU' else None,
            quartier        = str(localite).strip() if localite and str(localite).strip() != 'AU BUREAU' else None,
            adresse         = str(adresse).strip() if adresse and str(adresse).strip() != 'AU BUREAU' else None,
            superviseur     = str(superviseur).strip() if superviseur and str(superviseur).strip() != 'AU BUREAU' else None,
            gestionnaire    = str(gestionnaire).strip() if gestionnaire and str(gestionnaire).strip() != 'AU BUREAU' else None,
            teleconseillere = str(teleconseillere).strip() if teleconseillere and str(teleconseillere).strip() != 'AU BUREAU' else None,
            developpeur     = str(developpeur).strip() if developpeur and str(developpeur).strip() != 'AU BUREAU' else None,
            date_activation = date_activation,
            sim_au_bureau   = sim_bureau,
            nouvelle_creation = num_pdv_str in new_creation_map,
            notes           = ' | '.join(notes_parts) if notes_parts else None,

        )

        db.add(pdv)
        inseres += 1

        if inseres % 100 == 0:
            db.commit()
            print(f"  ✅ {inseres} PDVs importés...")

    except Exception as e:
        erreurs += 1
        print(f"  ⚠️ Erreur ligne {num_pdv}: {e}")

db.commit()
db.close()

print(f"\n🎉 Import terminé !")
print(f"  ✅ Insérés  : {inseres}")
print(f"  ⏭️  Ignorés  : {ignores}")
print(f"  ❌ Erreurs  : {erreurs}")
