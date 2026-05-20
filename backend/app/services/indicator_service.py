"""
Service Indicateurs - CRUD + calcul des scores + import.
=========================================================
"""
from __future__ import annotations
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from fastapi import HTTPException

from app.models.indicator import (
    Indicator, IndicatorVersion, IndicatorScore,
    IndicatorCategory, IndicatorMethod, IndicatorPeriod, IndicatorStatus,
)
from app.models.pdv import PDV
from app.models.user import User


# ─────────────────────────────────────────────────────────────────────────────
# Période courante
# ─────────────────────────────────────────────────────────────────────────────
def current_period_key(period: IndicatorPeriod) -> str:
    now = datetime.utcnow()
    if period == IndicatorPeriod.DAILY:
        return now.strftime("%Y-%m-%d")
    if period == IndicatorPeriod.WEEKLY:
        return f"{now.year}-W{now.isocalendar()[1]:02d}"
    return now.strftime("%Y-%m")


# ─────────────────────────────────────────────────────────────────────────────
# CRUD INDICATEURS
# ─────────────────────────────────────────────────────────────────────────────
def list_indicators(db: Session, status: Optional[IndicatorStatus] = None,
                    category: Optional[IndicatorCategory] = None) -> List[Indicator]:
    q = db.query(Indicator)
    if status: q = q.filter(Indicator.status == status)
    if category: q = q.filter(Indicator.category == category)
    return q.order_by(Indicator.name).all()


def get_indicator(db: Session, indicator_id: int) -> Indicator:
    i = db.query(Indicator).get(indicator_id)
    if not i:
        raise HTTPException(404, f"Indicateur {indicator_id} introuvable")
    return i


def create_indicator(db: Session, payload: dict, user_id: Optional[int] = None) -> Indicator:
    code = payload.get("code", "").strip().upper()
    if not code:
        raise HTTPException(400, "Code obligatoire")
    if db.query(Indicator).filter(Indicator.code == code).first():
        raise HTTPException(409, f"Code '{code}' déjà utilisé")
    indic = Indicator(
        code=code,
        name=payload.get("name") or code,
        description=payload.get("description"),
        category=IndicatorCategory(payload.get("category", "PRODUIT")),
        icon=payload.get("icon"),
        color=payload.get("color"),
        method=IndicatorMethod(payload.get("method", "MANUAL")),
        metric_field=payload.get("metric_field"),
        threshold_value=payload.get("threshold_value"),
        formula=payload.get("formula"),
        period=IndicatorPeriod(payload.get("period", "MONTHLY")),
        status=IndicatorStatus(payload.get("status", "ACTIVE")),
        target_rate_pct=payload.get("target_rate_pct"),
        weight=payload.get("weight", 1.0),
        created_by_id=user_id,
    )
    db.add(indic); db.flush()
    _save_version(db, indic, user_id, "Création initiale")
    db.commit(); db.refresh(indic)
    return indic


def update_indicator(db: Session, indicator_id: int, payload: dict, user_id: Optional[int] = None) -> Indicator:
    i = get_indicator(db, indicator_id)
    EDITABLE = {"name", "description", "category", "icon", "color",
                "method", "metric_field", "threshold_value", "formula",
                "period", "status", "target_rate_pct", "weight"}
    for k, v in payload.items():
        if k in EDITABLE and v is not None:
            if k == "category": v = IndicatorCategory(v)
            elif k == "method": v = IndicatorMethod(v)
            elif k == "period": v = IndicatorPeriod(v)
            elif k == "status": v = IndicatorStatus(v)
            setattr(i, k, v)
    i.updated_at = datetime.utcnow()
    _save_version(db, i, user_id, payload.get("change_note") or "Mise à jour")
    db.commit(); db.refresh(i)
    return i


def archive_indicator(db: Session, indicator_id: int) -> Indicator:
    i = get_indicator(db, indicator_id)
    i.status = IndicatorStatus.ARCHIVED
    db.commit(); db.refresh(i)
    return i


def _save_version(db: Session, indic: Indicator, user_id: Optional[int], note: str):
    v_no = (db.query(func.count(IndicatorVersion.id))
              .filter(IndicatorVersion.indicator_id == indic.id).scalar() or 0) + 1
    snap = {
        "code": indic.code, "name": indic.name, "method": indic.method.value,
        "metric_field": indic.metric_field, "threshold_value": indic.threshold_value,
        "formula": indic.formula, "period": indic.period.value, "status": indic.status.value,
        "target_rate_pct": indic.target_rate_pct, "weight": indic.weight,
    }
    db.add(IndicatorVersion(
        indicator_id=indic.id, version_no=v_no, snapshot=snap,
        changed_by_id=user_id, change_note=note,
    ))


# ─────────────────────────────────────────────────────────────────────────────
# SCORES — quels PDV font / ne font pas
# ─────────────────────────────────────────────────────────────────────────────
def set_score(db: Session, indicator_id: int, pdv_id: int, period_key: str,
              is_active: bool, raw_value: Optional[float] = None,
              source: str = "manual") -> IndicatorScore:
    s = db.query(IndicatorScore).filter(
        IndicatorScore.indicator_id == indicator_id,
        IndicatorScore.pdv_id == pdv_id,
        IndicatorScore.period_key == period_key,
    ).first()
    if s:
        s.is_active = is_active
        s.raw_value = raw_value
        s.source = source
        s.measured_at = datetime.utcnow()
    else:
        s = IndicatorScore(
            indicator_id=indicator_id, pdv_id=pdv_id, period_key=period_key,
            is_active=is_active, raw_value=raw_value, source=source,
        )
        db.add(s)
    db.commit(); db.refresh(s)
    return s


def get_pdvs_status(db: Session, indicator_id: int, period_key: Optional[str] = None,
                    active_only: Optional[bool] = None,
                    filters: Optional[dict] = None,
                    limit: int = 500) -> List[Dict[str, Any]]:
    """
    Renvoie les PDV avec leur statut sur l'indicateur pour la période.
    Version optimisée : JOIN SQL au lieu de 2 requêtes séparées + limit.
    """
    indic = get_indicator(db, indicator_id)
    period_key = period_key or current_period_key(indic.period)
    filters = filters or {}

    from sqlalchemy import outerjoin
    # JOIN PDV ← IndicatorScore (outerjoin pour avoir aussi les PDV sans score)
    score_alias = db.query(IndicatorScore).filter(
        IndicatorScore.indicator_id == indicator_id,
        IndicatorScore.period_key == period_key,
    ).subquery()

    q = db.query(PDV, score_alias.c.is_active, score_alias.c.raw_value, score_alias.c.measured_at)\
        .outerjoin(score_alias, PDV.id == score_alias.c.pdv_id)

    # Filtres
    if filters.get("quartier"):
        q = q.filter(PDV.quartier == filters["quartier"])
    if filters.get("search"):
        like = f"%{filters['search']}%"
        try:
            q = q.filter(or_(PDV.nom.ilike(like), PDV.numero_pdv.ilike(like)))
        except Exception:
            pass

    # Filtre actif/inactif directement dans SQL
    if active_only is True:
        q = q.filter(score_alias.c.is_active == True)
    elif active_only is False:
        q = q.filter(or_(score_alias.c.is_active == False, score_alias.c.is_active == None))

    out = []
    for p, is_active, raw_value, measured_at in q.limit(limit).all():
        out.append({
            "pdv_id": p.id,
            "numero_pdv": getattr(p, "numero_pdv", None),
            "nom": getattr(p, "nom", None),
            "telephone": getattr(p, "telephone", None),
            "quartier": getattr(p, "quartier", None),
            "zone": getattr(p, "zone", None),
            "sous_zone": getattr(p, "sous_zone", None),
            "statut_pdv": p.statut.value if getattr(p, "statut", None) and hasattr(p.statut, "value") else None,
            "lat": getattr(p, "latitude", None), "lng": getattr(p, "longitude", None),
            "is_active": bool(is_active) if is_active is not None else False,
            "raw_value": raw_value,
            "measured_at": measured_at.isoformat() if measured_at else None,
        })
    return out


def stats(db: Session, indicator_id: int, period_key: Optional[str] = None) -> Dict[str, Any]:
    indic = get_indicator(db, indicator_id)
    period_key = period_key or current_period_key(indic.period)
    total_pdvs = db.query(func.count(PDV.id)).scalar() or 0
    active = db.query(func.count(IndicatorScore.id)).filter(
        IndicatorScore.indicator_id == indicator_id,
        IndicatorScore.period_key == period_key,
        IndicatorScore.is_active == True,
    ).scalar() or 0
    return {
        "indicator_id": indicator_id, "period_key": period_key,
        "total_pdvs": total_pdvs, "active": active,
        "inactive": total_pdvs - active,
        "rate_pct": round(active / total_pdvs * 100, 1) if total_pdvs else 0,
        "target_pct": indic.target_rate_pct,
        "gap_to_target": round(indic.target_rate_pct - (active / total_pdvs * 100), 1)
                         if indic.target_rate_pct and total_pdvs else None,
    }


def stats_global(db: Session, period_key: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Version optimisée : 2 requêtes SQL au lieu de N×M.
    1. Compte total PDV
    2. Une seule requête GROUP BY pour les scores actifs par indicateur
    """
    indicators = db.query(Indicator).filter(Indicator.status == IndicatorStatus.ACTIVE).all()
    if not indicators:
        return []

    total_pdvs = db.query(func.count(PDV.id)).scalar() or 0

    # Période par défaut (mensuelle pour la vue globale)
    pk = period_key or datetime.utcnow().strftime("%Y-%m")

    # Une seule requête : nombre de scores actifs groupés par indicateur
    rows = db.query(
        IndicatorScore.indicator_id,
        func.count(IndicatorScore.id).label("active_count"),
    ).filter(
        IndicatorScore.period_key == pk,
        IndicatorScore.is_active == True,
        IndicatorScore.indicator_id.in_([i.id for i in indicators]),
    ).group_by(IndicatorScore.indicator_id).all()

    active_by_id = {r.indicator_id: r.active_count for r in rows}

    out = []
    for ind in indicators:
        active = active_by_id.get(ind.id, 0)
        rate = round(active / total_pdvs * 100, 1) if total_pdvs else 0
        gap = round(ind.target_rate_pct - rate, 1) if ind.target_rate_pct else None
        out.append({
            "indicator_id": ind.id,
            "code": ind.code, "name": ind.name, "icon": ind.icon, "color": ind.color,
            "category": ind.category.value,
            "period_key": pk,
            "total_pdvs": total_pdvs,
            "active": active,
            "inactive": total_pdvs - active,
            "rate_pct": rate,
            "target_pct": ind.target_rate_pct,
            "gap_to_target": gap,
        })
    return out


def evolution(db: Session, indicator_id: int, n_periods: int = 12) -> List[Dict[str, Any]]:
    """Retourne l'évolution du taux sur les N dernières périodes."""
    indic = get_indicator(db, indicator_id)
    now = datetime.utcnow()
    keys = []
    for i in range(n_periods - 1, -1, -1):
        if indic.period == IndicatorPeriod.MONTHLY:
            d = (now.replace(day=1) - timedelta(days=30 * i))
            keys.append(d.strftime("%Y-%m"))
        elif indic.period == IndicatorPeriod.WEEKLY:
            d = now - timedelta(weeks=i)
            keys.append(f"{d.year}-W{d.isocalendar()[1]:02d}")
        else:
            keys.append((now - timedelta(days=i)).strftime("%Y-%m-%d"))
    keys = list(dict.fromkeys(keys))  # uniques en gardant l'ordre

    total_pdvs = db.query(func.count(PDV.id)).scalar() or 0
    out = []
    for k in keys:
        active = db.query(func.count(IndicatorScore.id)).filter(
            IndicatorScore.indicator_id == indicator_id,
            IndicatorScore.period_key == k,
            IndicatorScore.is_active == True,
        ).scalar() or 0
        out.append({
            "period_key": k, "active": active, "total": total_pdvs,
            "rate_pct": round(active / total_pdvs * 100, 1) if total_pdvs else 0,
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
# IMPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def import_xlsx(db: Session, indicator_id: int, period_key: str,
                file_bytes: bytes,
                pdv_col: str = "numero_pdv",
                value_col: Optional[str] = "valeur",
                active_col: Optional[str] = None,
                source: str = "import_xlsx") -> Dict[str, Any]:
    """
    Importer un Excel pour un indicateur :
      - pdv_col   : colonne contenant le n° de PDV
      - value_col : colonne contenant la métrique (CA, etc.)
      - active_col : colonne booléenne (oui/non)
    Si l'indicateur est THRESHOLD, on calcule is_active à partir de value_col.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    indic = get_indicator(db, indicator_id)
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    h = {h.lower(): i for i, h in enumerate(headers)}

    if pdv_col.lower() not in h:
        raise HTTPException(400, f"Colonne '{pdv_col}' introuvable. Trouvées: {headers}")

    idx_pdv = h[pdv_col.lower()]
    idx_val = h.get(value_col.lower()) if value_col else None
    idx_act = h.get(active_col.lower()) if active_col else None

    # Préparer index PDV par numéro
    pdv_index = {str(p.numero_pdv): p.id for p in db.query(PDV).filter(PDV.numero_pdv.isnot(None)).all()}

    created, updated, skipped = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx_pdv] is None:
            skipped += 1; continue
        num = str(row[idx_pdv]).strip()
        pdv_id = pdv_index.get(num)
        if not pdv_id:
            skipped += 1; continue

        raw_value = float(row[idx_val]) if (idx_val is not None and row[idx_val] not in (None, "")) else None
        is_active = None
        if idx_act is not None:
            cell = str(row[idx_act] or "").lower()
            is_active = cell in ("oui", "yes", "true", "1", "actif", "x", "✓")
        elif indic.method == IndicatorMethod.THRESHOLD and raw_value is not None and indic.threshold_value is not None:
            is_active = raw_value >= indic.threshold_value
        elif raw_value is not None:
            is_active = raw_value > 0
        else:
            is_active = False

        existing = db.query(IndicatorScore).filter(
            IndicatorScore.indicator_id == indicator_id,
            IndicatorScore.pdv_id == pdv_id,
            IndicatorScore.period_key == period_key,
        ).first()
        if existing:
            existing.is_active = bool(is_active)
            existing.raw_value = raw_value
            existing.source = source
            existing.measured_at = datetime.utcnow()
            updated += 1
        else:
            db.add(IndicatorScore(
                indicator_id=indicator_id, pdv_id=pdv_id, period_key=period_key,
                is_active=bool(is_active), raw_value=raw_value, source=source,
            ))
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}
