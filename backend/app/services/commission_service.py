"""
Service Commissions Réseau.
============================
Calcul, import Excel, reporting par type/zone/période.

Règle universelle :
  montant_reseau = brut × 30%
  montant_pdv    = brut × 70%
  gere_reversement = True si KIOSQUE ou RS (PDG reçoit les 70% et doit les reverser)
"""
from __future__ import annotations
from datetime import datetime, timedelta
from io import BytesIO
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from fastapi import HTTPException

from app.models.commission import (
    CommissionEntry, CommissionImport, PDVType, ReversementStatus,
    TYPE_GERE_REVERSEMENT, TAUX_RESEAU, TAUX_PDV,
)
from app.models.pdv import PDV


# ─────────────────────────────────────────────────────────────────────────────
# Import Excel
# ─────────────────────────────────────────────────────────────────────────────
def import_xlsx(
    db: Session,
    file_bytes: bytes,
    filename: str,
    period_key: str,
    period_type: str = "MONTHLY",
    col_mapping: Optional[Dict[str, str]] = None,
    user_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Importe un Excel de commissions Orange.

    Colonnes attendues (noms configurables via col_mapping) :
      - numero_pdv   : N° du PDV
      - pdv_nom      : Nom du PDV
      - pdv_type     : Type (RNS, RSF, RS, KIOSQUE)
      - montant_brut : Montant 100% reçu d'Orange
      - quartier     : Quartier (optionnel)
      - zone         : Zone (optionnel)
      - gestionnaire : Gestionnaire (optionnel)
      - superviseur  : Superviseur (optionnel)
    """
    from openpyxl import load_workbook

    # Mapping par défaut des colonnes
    default_map = {
        "numero_pdv":   "numero_pdv",
        "pdv_nom":      "pdv_nom",
        "pdv_type":     "pdv_type",
        "montant_brut": "montant_brut",
        "quartier":     "quartier",
        "zone":         "zone",
        "sous_zone":    "sous_zone",
        "gestionnaire": "gestionnaire",
        "superviseur":  "superviseur",
    }
    if col_mapping:
        default_map.update(col_mapping)

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    h = {v: i for i, v in enumerate(headers)}

    def get_col(field: str):
        col_name = default_map.get(field, field).lower()
        return h.get(col_name)

    required = ["numero_pdv", "pdv_type", "montant_brut"]
    for r in required:
        if get_col(r) is None:
            raise HTTPException(
                400,
                f"Colonne obligatoire '{default_map[r]}' introuvable. "
                f"Colonnes trouvées : {list(h.keys())}"
            )

    # Index PDV
    pdv_index = {
        str(p.numero_pdv).strip(): p.id
        for p in db.query(PDV).filter(PDV.numero_pdv.isnot(None)).all()
    }

    created, updated, skipped = 0, 0, 0
    now = datetime.utcnow()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[get_col("numero_pdv")] is None:
            skipped += 1; continue

        numero = str(row[get_col("numero_pdv")]).strip()
        type_raw = str(row[get_col("pdv_type")] or "").strip().upper()
        brut_raw = row[get_col("montant_brut")]

        # Validation type
        try:
            pdv_type = PDVType(type_raw)
        except ValueError:
            skipped += 1; continue

        # Validation montant
        try:
            montant_brut = float(str(brut_raw).replace(" ", "").replace(",", "."))
            if montant_brut < 0:
                skipped += 1; continue
        except (TypeError, ValueError):
            skipped += 1; continue

        # Calcul des parts
        montant_reseau = round(montant_brut * TAUX_RESEAU, 2)
        montant_pdv    = round(montant_brut * TAUX_PDV, 2)
        gere_rev = TYPE_GERE_REVERSEMENT[pdv_type]

        # Champs optionnels
        def g(f):
            idx = get_col(f)
            return str(row[idx]).strip() if idx is not None and row[idx] not in (None, "") else None

        quartier    = g("quartier")
        zone        = g("zone")
        sous_zone   = g("sous_zone")
        gestionnaire = g("gestionnaire")
        superviseur  = g("superviseur")
        pdv_nom      = g("pdv_nom")
        pdv_id       = pdv_index.get(numero)

        # Upsert
        existing = db.query(CommissionEntry).filter(
            CommissionEntry.pdv_numero == numero,
            CommissionEntry.period_key == period_key,
        ).first()

        if existing:
            existing.pdv_type      = pdv_type
            existing.montant_brut  = montant_brut
            existing.montant_reseau = montant_reseau
            existing.montant_pdv   = montant_pdv
            existing.gere_reversement = gere_rev
            existing.quartier      = quartier or existing.quartier
            existing.zone          = zone or existing.zone
            existing.sous_zone     = sous_zone or existing.sous_zone
            existing.gestionnaire  = gestionnaire or existing.gestionnaire
            existing.superviseur   = superviseur or existing.superviseur
            existing.imported_at   = now
            if existing.reversement_status == ReversementStatus.NON_APPLICABLE and gere_rev:
                existing.reversement_status = ReversementStatus.EN_ATTENTE
            updated += 1
        else:
            db.add(CommissionEntry(
                pdv_id=pdv_id, pdv_numero=numero, pdv_nom=pdv_nom,
                pdv_type=pdv_type, quartier=quartier, zone=zone,
                sous_zone=sous_zone, gestionnaire=gestionnaire,
                superviseur=superviseur, period_key=period_key,
                period_type=period_type,
                montant_brut=montant_brut,
                montant_reseau=montant_reseau,
                montant_pdv=montant_pdv,
                gere_reversement=gere_rev,
                reversement_status=(
                    ReversementStatus.EN_ATTENTE if gere_rev
                    else ReversementStatus.NON_APPLICABLE
                ),
                source="import_xlsx",
                imported_by_id=user_id,
            ))
            created += 1

    db.commit()

    # Trace de l'import
    db.add(CommissionImport(
        filename=filename, period_key=period_key, period_type=period_type,
        imported_by_id=user_id, n_created=created, n_updated=updated, n_skipped=skipped,
    ))
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "period_key": period_key}


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD GLOBAL
# ─────────────────────────────────────────────────────────────────────────────
def dashboard(db: Session, period_key: str, pdv_type: Optional[PDVType] = None,
              superviseur: Optional[str] = None, gestionnaire: Optional[str] = None,
              zone: Optional[str] = None) -> Dict[str, Any]:
    """
    Vue synthétique pour une période.

    Logique Orange Mali :
      ─ RNS / RSF : Orange verse 100% au PDV directement.
                    Le réseau reçoit SEULEMENT ses 30% (montant_reseau).
                    → gere_reversement = False
      ─ RS / KIOSQUE : Orange verse 100% au PDG.
                    Le PDG garde ses 30% ET détient les 70% du PDV en transit.
                    → gere_reversement = True

    Trois niveaux de lecture du "montant réseau" :
      1. commission_brute   = Σ(montant_reseau) tous types       ← ce que le réseau DOIT garder
      2. montant_en_transit = Σ(montant_pdv) RS+KIOSQUE          ← 70% reçus, à reverser
      3. commission_nette   = commission_brute + reste_a_reverser ← ce que le PDG a EN CAISSE maintenant
                            = commission_brute si tout est reversé
    """
    q = db.query(CommissionEntry).filter(CommissionEntry.period_key == period_key)
    if pdv_type:
        q = q.filter(CommissionEntry.pdv_type == pdv_type)
    if superviseur:
        q = q.filter(CommissionEntry.superviseur.ilike(f"%{superviseur}%"))
    if gestionnaire:
        q = q.filter(CommissionEntry.gestionnaire.ilike(f"%{gestionnaire}%"))
    if zone:
        q = q.filter(CommissionEntry.zone == zone)

    entries = q.all()
    if not entries:
        return _empty_dashboard(period_key)

    # ── Séparation RNS/RSF vs RS/KIOSQUE ──────────────────────────────────────
    ents_directs = [e for e in entries if not e.gere_reversement]   # RNS + RSF
    ents_geres   = [e for e in entries if e.gere_reversement]       # RS + KIOSQUE

    total_brut   = sum(e.montant_brut for e in entries)
    total_reseau = sum(e.montant_reseau for e in entries)   # Σ 30% tous types
    total_pdv    = sum(e.montant_pdv for e in entries)      # Σ 70% tous types

    # ── Commission BRUTE réseau (30% pour tous les PDV) ──────────────────────
    commission_brute_rns_rsf    = sum(e.montant_reseau for e in ents_directs)
    commission_brute_rs_kiosque = sum(e.montant_reseau for e in ents_geres)
    commission_brute_total      = round(total_reseau, 2)   # = commission_brute_rns_rsf + commission_brute_rs_kiosque

    # ── Montant en transit (70% RS+KIOSQUE que le PDG détient) ───────────────
    montant_en_transit = sum(e.montant_pdv for e in ents_geres)     # à reverser
    total_reverse      = sum(e.montant_reverse or 0 for e in ents_geres)
    total_reste        = montant_en_transit - total_reverse          # reste dû aux PDV

    # ── Commission NETTE réseau (ce que le PDG a réellement en caisse) ────────
    # = ses 30% + les 70% RS/KIOSQUE pas encore reversés
    commission_nette = commission_brute_total + total_reste

    # ── Montant total reçu par le PDG en trésorerie ───────────────────────────
    # PDG reçoit 30% de toutes les commissions brutes
    # commission_brute_total = total_reseau = Σ(montant_reseau) = Σ(brut * 30%)
    montant_recu_pdg = round(commission_brute_total, 2)

    # ── Ventilation par type ───────────────────────────────────────────────────
    by_type = {}
    for t in PDVType:
        ents = [e for e in entries if e.pdv_type == t]
        if not ents: continue
        gere = TYPE_GERE_REVERSEMENT[t]
        brut_t    = sum(e.montant_brut for e in ents)
        reseau_t  = sum(e.montant_reseau for e in ents)
        pdv_t     = sum(e.montant_pdv for e in ents)
        reverse_t = sum(e.montant_reverse or 0 for e in ents) if gere else 0
        reste_t   = (pdv_t - reverse_t) if gere else 0
        by_type[t.value] = {
            "type": t.value,
            "n_pdv": len(ents),
            "brut":   round(brut_t, 2),
            "reseau": round(reseau_t, 2),    # 30% (part réseau)
            "pdv":    round(pdv_t, 2),        # 70% (part PDV)
            "gere_reversement": gere,
            # Spécifique RS/KIOSQUE
            "en_transit":       round(pdv_t, 2) if gere else 0,
            "deja_reverse":     round(reverse_t, 2),
            "reste_a_reverser": round(reste_t, 2),
            # Commission nette pour ce type
            "commission_nette": round(reseau_t + reste_t, 2),
        }

    # ── Reversements (KIOSQUE + RS uniquement) ────────────────────────────────
    by_rev_status: Dict[str, int] = {}
    for e in ents_geres:
        s = e.reversement_status.value
        by_rev_status[s] = by_rev_status.get(s, 0) + 1

    # ── Ventilation par quartier (top 15) ─────────────────────────────────────
    by_quartier: Dict[str, Any] = {}
    for e in entries:
        q_key = e.quartier or "Non renseigné"
        if q_key not in by_quartier:
            by_quartier[q_key] = {"quartier": q_key, "n_pdv": 0, "brut": 0, "reseau": 0, "pdv": 0, "commission_nette": 0}
        by_quartier[q_key]["n_pdv"]  += 1
        by_quartier[q_key]["brut"]   += e.montant_brut
        by_quartier[q_key]["reseau"] += e.montant_reseau
        by_quartier[q_key]["pdv"]    += e.montant_pdv
        reste_e = (e.montant_pdv - (e.montant_reverse or 0)) if e.gere_reversement else 0
        by_quartier[q_key]["commission_nette"] += e.montant_reseau + reste_e
    top_quartiers = sorted(by_quartier.values(), key=lambda x: x["brut"], reverse=True)[:15]
    for qq in top_quartiers:
        qq["brut"]             = round(qq["brut"], 2)
        qq["reseau"]           = round(qq["reseau"], 2)
        qq["pdv"]              = round(qq["pdv"], 2)
        qq["commission_nette"] = round(qq["commission_nette"], 2)

    # ── Ventilation par zone ──────────────────────────────────────────────────
    by_zone: Dict[str, Any] = {}
    for e in entries:
        z_key = e.zone or "Non renseigné"
        if z_key not in by_zone:
            by_zone[z_key] = {"zone": z_key, "n_pdv": 0, "brut": 0, "reseau": 0, "pdv": 0, "commission_nette": 0}
        by_zone[z_key]["n_pdv"]  += 1
        by_zone[z_key]["brut"]   += e.montant_brut
        by_zone[z_key]["reseau"] += e.montant_reseau
        by_zone[z_key]["pdv"]    += e.montant_pdv
        reste_e = (e.montant_pdv - (e.montant_reverse or 0)) if e.gere_reversement else 0
        by_zone[z_key]["commission_nette"] += e.montant_reseau + reste_e
    top_zones = sorted(by_zone.values(), key=lambda x: x["brut"], reverse=True)
    for zz in top_zones:
        zz["brut"]             = round(zz["brut"], 2)
        zz["reseau"]           = round(zz["reseau"], 2)
        zz["pdv"]              = round(zz["pdv"], 2)
        zz["commission_nette"] = round(zz["commission_nette"], 2)

    return {
        "period_key":   period_key,
        "n_pdv_total":  len(entries),
        "n_pdv_directs": len(ents_directs),    # RNS + RSF
        "n_pdv_geres":   len(ents_geres),      # RS + KIOSQUE

        # ── Montants globaux bruts ─────────────────────────────────────────
        "total_brut":  round(total_brut, 2),   # 100% tous PDV
        "total_reseau": round(total_reseau, 2), # Σ 30% = commission_brute_total
        "total_pdv":   round(total_pdv, 2),     # Σ 70%
        "taux_reseau": TAUX_RESEAU * 100,
        "taux_pdv":    TAUX_PDV * 100,
 
        # ── Lecture réseau (3 niveaux) ─────────────────────────────────────
        "commission_brute": {
            "total":      round(commission_brute_total, 2),
            "rns_rsf":    round(commission_brute_rns_rsf, 2),     # 30% reçus directement
            "rs_kiosque": round(commission_brute_rs_kiosque, 2),  # 30% sur les 100% reçus
        },
        "montant_en_transit": {
            "total":          round(montant_en_transit, 2),  # 70% RS+KIOSQUE reçus par PDG
            "deja_reverse":   round(total_reverse, 2),       # déjà payés aux PDV
            "reste_a_payer":  round(total_reste, 2),         # encore dû aux PDV RS/KIOSQUE
            "taux_reversement": round(
                (total_reverse / montant_en_transit * 100) if montant_en_transit > 0 else 0, 1
            ),
        },
        "commission_nette": round(commission_nette, 2),
        # = commission_brute_total + reste_a_payer
        # = ce que le PDG a réellement EN CAISSE maintenant

        "montant_recu_pdg": round(montant_recu_pdg, 2),
        # = 30%(RNS+RSF) + 100%(RS+KIOSQUE) = total reçu en trésorerie

        # ── Ventilations ──────────────────────────────────────────────────
        "by_type":     list(by_type.values()),
        "by_quartier": top_quartiers,
        "by_zone":     top_zones,

        # ── Reversements détail (KIOSQUE + RS) ────────────────────────────
        "reversements": {
            "total_a_reverser":  round(montant_en_transit, 2),
            "total_reverse":     round(total_reverse, 2),
            "total_reste":       round(total_reste, 2),
            "n_pdv_concernes":   len(ents_geres),
            "by_status":         by_rev_status,
        },
    }


def _empty_dashboard(period_key: str) -> Dict[str, Any]:
    empty_transit = {"total": 0, "deja_reverse": 0, "reste_a_payer": 0, "taux_reversement": 0}
    empty_brute   = {"total": 0, "rns_rsf": 0, "rs_kiosque": 0}
    return {
        "period_key": period_key, "n_pdv_total": 0,
        "n_pdv_directs": 0, "n_pdv_geres": 0,
        "total_brut": 0, "total_reseau": 0, "total_pdv": 0,
        "taux_reseau": 30, "taux_pdv": 70,
        "commission_brute":    empty_brute,
        "montant_en_transit":  empty_transit,
        "commission_nette":    0,
        "montant_recu_pdg":    0,
        "by_type": [], "by_quartier": [], "by_zone": [],
        "reversements": {"total_a_reverser": 0, "total_reverse": 0, "total_reste": 0,
                         "n_pdv_concernes": 0, "by_status": {}},
    }


# ─────────────────────────────────────────────────────────────────────────────
# LISTE DÉTAILLÉE DES PDV
# ─────────────────────────────────────────────────────────────────────────────
def list_entries(
    db: Session, period_key: str,
    pdv_type: Optional[PDVType] = None,
    quartier: Optional[str] = None,
    zone: Optional[str] = None,
    reversement_status: Optional[ReversementStatus] = None,
    search: Optional[str] = None,
    gere_reversement: Optional[bool] = None,
    skip: int = 0, limit: int = 200,
    superviseur: Optional[str] = None,
    gestionnaire: Optional[str] = None,
) -> List[Dict[str, Any]]:
    q = db.query(CommissionEntry).filter(CommissionEntry.period_key == period_key)
    if pdv_type: q = q.filter(CommissionEntry.pdv_type == pdv_type)
    if quartier: q = q.filter(CommissionEntry.quartier == quartier)
    if zone: q = q.filter(CommissionEntry.zone == zone)
    if superviseur: q = q.filter(CommissionEntry.superviseur.ilike(f"%{superviseur}%"))
    if gestionnaire: q = q.filter(CommissionEntry.gestionnaire.ilike(f"%{gestionnaire}%"))
    if reversement_status: q = q.filter(CommissionEntry.reversement_status == reversement_status)
    if gere_reversement is not None: q = q.filter(CommissionEntry.gere_reversement == gere_reversement)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(
            CommissionEntry.pdv_numero.ilike(like),
            CommissionEntry.pdv_nom.ilike(like),
            CommissionEntry.quartier.ilike(like),
        ))
    entries = q.order_by(CommissionEntry.montant_brut.desc()).offset(skip).limit(limit).all()
    return [_entry_to_dict(e) for e in entries]


def _entry_to_dict(e: CommissionEntry) -> Dict[str, Any]:
    montant_pdv    = e.montant_pdv or 0
    montant_reverse = e.montant_reverse or 0
    reste = round(montant_pdv - montant_reverse, 2) if e.gere_reversement else 0

    # Commission NETTE pour ce PDV :
    # ─ RNS/RSF    : juste les 30% (les 70% sont partis directement au PDV via Orange)
    # ─ RS/KIOSQUE : 30% + les 70% pas encore reversés (encore en caisse du PDG)
    commission_nette = round((e.montant_reseau or 0) + reste, 2)

    return {
        "id": e.id, "pdv_numero": e.pdv_numero, "pdv_nom": e.pdv_nom,
        "pdv_type": e.pdv_type.value, "quartier": e.quartier,
        "zone": e.zone, "sous_zone": e.sous_zone,
        "gestionnaire": e.gestionnaire, "superviseur": e.superviseur,
        "period_key": e.period_key,
        # Montants bruts
        "montant_brut":   e.montant_brut,
        "montant_reseau": e.montant_reseau,   # 30% — part réseau
        "montant_pdv":    montant_pdv,         # 70% — part PDV
        # Commission nette réseau (ce que le PDG garde vraiment pour ce PDV)
        "commission_nette": commission_nette,
        # Reversement (RS/KIOSQUE uniquement)
        "gere_reversement": e.gere_reversement,
        "reversement_status": e.reversement_status.value,
        "montant_reverse": montant_reverse,
        "montant_reste_a_reverser": reste,
        "date_reversement": e.date_reversement.isoformat() if e.date_reversement else None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# ÉVOLUTION MULTI-PÉRIODES
# ─────────────────────────────────────────────────────────────────────────────
def evolution(db: Session, n_periods: int = 6,
              pdv_type: Optional[PDVType] = None,
              superviseur: Optional[str] = None, gestionnaire: Optional[str] = None,
              zone: Optional[str] = None) -> List[Dict[str, Any]]:
    """Tendance sur les N dernières périodes mensuelles."""
    now = datetime.utcnow()
    keys = [(now - timedelta(days=30 * i)).strftime("%Y-%m")
            for i in range(n_periods - 1, -1, -1)]

    out = []
    for pk in keys:
        q = db.query(
            func.count(CommissionEntry.id).label("n"),
            func.sum(CommissionEntry.montant_brut).label("brut"),
            func.sum(CommissionEntry.montant_reseau).label("reseau"),
            func.sum(CommissionEntry.montant_pdv).label("pdv_total"),
        ).filter(CommissionEntry.period_key == pk)
        if pdv_type: q = q.filter(CommissionEntry.pdv_type == pdv_type)
        if superviseur: q = q.filter(CommissionEntry.superviseur.ilike(f"%{superviseur}%"))
        if gestionnaire: q = q.filter(CommissionEntry.gestionnaire.ilike(f"%{gestionnaire}%"))
        if zone: q = q.filter(CommissionEntry.zone == zone)
        r = q.first()
        out.append({
            "period_key": pk,
            "n_pdv": r.n or 0,
            "brut":   round(r.brut or 0, 2),
            "reseau": round(r.reseau or 0, 2),
            "pdv":    round(r.pdv_total or 0, 2),
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
# TOP PDV
# ─────────────────────────────────────────────────────────────────────────────
def top_pdvs(db: Session, period_key: str, n: int = 20,
             pdv_type: Optional[PDVType] = None,
             superviseur: Optional[str] = None, gestionnaire: Optional[str] = None,
             zone: Optional[str] = None) -> List[Dict[str, Any]]:
    q = db.query(CommissionEntry).filter(CommissionEntry.period_key == period_key)
    if pdv_type: q = q.filter(CommissionEntry.pdv_type == pdv_type)
    if superviseur: q = q.filter(CommissionEntry.superviseur.ilike(f"%{superviseur}%"))
    if gestionnaire: q = q.filter(CommissionEntry.gestionnaire.ilike(f"%{gestionnaire}%"))
    if zone: q = q.filter(CommissionEntry.zone == zone)
    entries = q.order_by(CommissionEntry.montant_brut.desc()).limit(n).all()
    return [_entry_to_dict(e) for e in entries]


# ─────────────────────────────────────────────────────────────────────────────
# PÉRIODES DISPONIBLES
# ─────────────────────────────────────────────────────────────────────────────
def available_periods(db: Session) -> List[str]:
    rows = db.query(CommissionEntry.period_key).distinct()\
        .order_by(CommissionEntry.period_key.desc()).all()
    return [r[0] for r in rows]


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def export_xlsx(db: Session, period_key: str,
                pdv_type: Optional[PDVType] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    entries = list_entries(db, period_key, pdv_type, limit=5000)
    wb = Workbook(); ws = wb.active
    ws.title = f"Commissions {period_key}"

    headers = [
        "N° PDV", "Nom PDV", "Type", "Quartier", "Zone",
        "Brut (100%)", "Réseau (30%)", "PDV (70%)",
        "Reversement géré", "Statut reversement", "Montant reversé", "Reste à reverser",
    ]
    fill = PatternFill("solid", fgColor="FF6900")
    font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center")

    for e in entries:
        ws.append([
            e["pdv_numero"], e["pdv_nom"], e["pdv_type"],
            e["quartier"], e["zone"],
            e["montant_brut"], e["montant_reseau"], e["montant_pdv"],
            "Oui" if e["gere_reversement"] else "Non",
            e["reversement_status"],
            e["montant_reverse"], e["montant_reste_a_reverser"],
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value)) for c in col if c.value), default=10) + 2

    from io import BytesIO
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
