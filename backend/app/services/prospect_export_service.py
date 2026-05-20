"""
Service Export.
================
- export_xlsx : exporte les prospects (avec filtres) en Excel
"""
from io import BytesIO
from typing import Optional, List
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.prospect import Prospect, ProspectStatus


HEADERS = [
    "Référence", "Statut", "Nom", "Prénom",
    "Téléphone principal", "Téléphone secondaire",
    "Quartier", "Adresse",
    "Pièce ID", "N° pièce",
    "Faisait OM", "CA mensuel OM", "Commission OM",
    "Capital démarrage", "Source financement",
    "Latitude", "Longitude",
    "Type local", "Fréquentation", "Concurrents",
    "Soumis le", "Visite assignée le", "Décision dev le",
    "Décision RC le", "Puce attribuée le", "Activée le",
    "N° Puce", "Notes",
]


def export_xlsx(db: Session, status_filter: Optional[ProspectStatus] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospection OM"

    # En-têtes stylés
    header_fill = PatternFill(start_color="FF6900", end_color="FF6900", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    q = db.query(Prospect)
    if status_filter:
        q = q.filter(Prospect.status == status_filter)
    rows = q.order_by(Prospect.created_at.desc()).all()

    def fmt_dt(dt): return dt.strftime("%Y-%m-%d %H:%M") if dt else ""
    for p in rows:
        ws.append([
            p.reference, p.status.value if hasattr(p.status, "value") else str(p.status),
            p.nom, p.prenom,
            p.telephone_principal, p.telephone_secondaire or "",
            p.quartier or "", p.adresse or "",
            p.piece_identite_type.value if p.piece_identite_type and hasattr(p.piece_identite_type, "value") else (str(p.piece_identite_type) if p.piece_identite_type else ""),
            p.piece_identite_numero or "",
            "Oui" if p.fait_om else "Non",
            p.om_ca_mensuel or "", p.om_commission_mensuelle or "",
            p.capital_demarrage or "", p.source_financement or "",
            p.latitude or "", p.longitude or "",
            p.type_local.value if p.type_local and hasattr(p.type_local, "value") else (str(p.type_local) if p.type_local else ""),
            p.frequentation.value if p.frequentation and hasattr(p.frequentation, "value") else (str(p.frequentation) if p.frequentation else ""),
            ", ".join(p.concurrents or []),
            fmt_dt(p.submitted_at), fmt_dt(p.visit_assigned_at), fmt_dt(p.dev_decision_at),
            fmt_dt(p.rc_decision_at), fmt_dt(p.puce_assigned_at), fmt_dt(p.activated_at),
            p.puce_numero or "", p.notes or "",
        ])

    # Largeurs auto basique
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()
