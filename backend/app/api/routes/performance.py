from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Form
from sqlalchemy.orm import Session
from typing import List, Optional
from app.core.database import get_db
from app.models.pdv import PDV
from app.models.performance import WeeklyPerformance, MonthlyPerformance
from app.models.user import User
from app.api.routes.auth import require_admin
from pydantic import BaseModel
from datetime import datetime
import io
import csv
import openpyxl

router = APIRouter()

class PerformanceBase(BaseModel):
    pdv_id: int
    ca: float = 0.0
    nb_operations: int = 0
    nb_depots: int = 0
    montant_depots: float = 0.0
    nb_retraits: int = 0
    montant_retraits: float = 0.0
    est_actif: bool = False

class WeeklyPerformanceCreate(PerformanceBase):
    annee: int
    semaine: int

class MonthlyPerformanceCreate(PerformanceBase):
    annee: int
    mois: int

class WeeklyPerformanceResponse(WeeklyPerformanceCreate):
    id: int
    ca_semaine_precedente: float
    taux_variation: float
    created_at: datetime

    class Config:
        from_attributes = True

class MonthlyPerformanceResponse(MonthlyPerformanceCreate):
    id: int
    ca_mois_precedent: float
    taux_variation: float
    created_at: datetime

    class Config:
        from_attributes = True

def _parse_file_to_rows(contents: bytes, filename: str):
    """Parse Excel or CSV file and return list of dicts"""
    fname = filename.lower()
    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        # Chercher la feuille SOURCE en priorité, sinon la première feuille active
        ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == 'SOURCE':
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active
        headers = None
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c is not None else '' for c in row]
            else:
                if all(c is None for c in row):
                    continue
                rows.append({headers[j]: row[j] for j in range(len(headers))})
        return rows
    else:
        # CSV
        text = contents.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

def _normalize_row(row: dict) -> dict:
    """Normalise les colonnes du format OMY natif vers le format standard attendu.
    Accepte le format natif exact: PDV, Année, Semaine, CA, NBRE OPERATIONS,
    Nbre dépôt, Montant Dépôt, Nbre Retrait, Montant Retrait
    ainsi que les variantes et le format template standard."""
    mapping = {
        # PDV
        'PDV': 'numero_pdv',
        'pdv': 'numero_pdv',
        'numero_pdv': 'numero_pdv',
        'numero': 'numero_pdv',
        # Année
        'Année': 'annee',
        'Annee': 'annee',
        'ANNEE': 'annee',
        'annee': 'annee',
        'année': 'annee',
        # Semaine
        'Semaine': 'semaine',
        'SEMAINE': 'semaine',
        'semaine': 'semaine',
        # Mois
        'Mois': 'mois',
        'MOIS': 'mois',
        'mois': 'mois',
        # CA
        'CA': 'ca',
        'ca': 'ca',
        # Opérations — format exact fichier: 'NBRE OPERATIONS'
        'NBRE OPERATIONS': 'nb_operations',
        'Nbre Operations': 'nb_operations',
        'Opération': 'nb_operations',
        'Operation': 'nb_operations',
        'OPERATION': 'nb_operations',
        'nb_operations': 'nb_operations',
        # Dépôts — format exact fichier: 'Nbre dépôt' et 'Montant Dépôt'
        'Nbre dépôt': 'nb_depots',
        'Nbre depot': 'nb_depots',
        'NBRE DEPOT': 'nb_depots',
        'nb_depots': 'nb_depots',
        'Montant Dépôt': 'montant_depots',
        'Montant depot': 'montant_depots',
        'MONTANT DEPOT': 'montant_depots',
        'Dépôt': 'montant_depots',
        'Depot': 'montant_depots',
        'montant_depots': 'montant_depots',
        # Retraits — format exact fichier: 'Nbre Retrait' et 'Montant Retrait'
        'Nbre Retrait': 'nb_retraits',
        'NBRE RETRAIT': 'nb_retraits',
        'nb_retraits': 'nb_retraits',
        'Montant Retrait': 'montant_retraits',
        'MONTANT RETRAIT': 'montant_retraits',
        'Retrait': 'montant_retraits',
        'montant_retraits': 'montant_retraits',
        # Actif
        'est_actif': 'est_actif',
    }
    normalized = {}
    for k, v in row.items():
        key = k.strip() if k else ''
        std_key = mapping.get(key, key)
        if std_key not in normalized:
            normalized[std_key] = v
    return normalized

def _parse_semaine(val) -> int:
    """Convertit 'S35', 'S 6', 35 → 35"""
    if val is None:
        return 0
    s = str(val).strip().upper().replace(' ', '')
    if s.startswith('S'):
        s = s[1:]
    try:
        return int(float(s))
    except:
        return 0

MOIS_MAP_STR = {
    'JANVIER': 1, 'FEVRIER': 2, 'FÉVRIER': 2, 'MARS': 3,
    'AVRIL': 4, 'MAI': 5, 'JUIN': 6, 'JUILLET': 7,
    'AOUT': 8, 'AOÛT': 8, 'SEPTEMBRE': 9, 'OCTOBRE': 10,
    'NOVEMBRE': 11, 'DECEMBRE': 12, 'DÉCEMBRE': 12
}

def _parse_mois(val) -> int:
    """Convertit 'Janvier', 1, '01' → 1"""
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    s = str(val).strip().upper()
    if s in MOIS_MAP_STR:
        return MOIS_MAP_STR[s]
    try:
        return int(float(s))
    except:
        return 0

def _safe_float(val, default=0.0):
    try:
        if val is None or str(val).strip() in ('', 'None', 'N/A', '-'):
            return default
        return float(str(val).replace(' ', '').replace(',', '.'))
    except:
        return default

def _safe_int(val, default=0):
    try:
        if val is None or str(val).strip() in ('', 'None', 'N/A', '-'):
            return default
        return int(float(str(val).replace(' ', '').replace(',', '.')))
    except:
        return default

def _safe_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ('true', '1', 'oui', 'yes', 'actif', 'vrai')

@router.post("/performance/weekly")
def bulk_import_weekly(
    file: UploadFile = File(...),
    indicateur: Optional[str] = Query(default='OMY'),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Bulk import weekly performance data (Excel ou CSV, numero_pdv)"""
    try:
        contents = file.file.read()
        rows = _parse_file_to_rows(contents, file.filename)

        # Construire un cache numero_pdv -> pdv_id
        all_pdvs = db.query(PDV.id, PDV.numero_pdv).all()
        pdv_map = {str(p.numero_pdv): p.id for p in all_pdvs}

        created_count = 0
        updated_count = 0
        errors = []
        total = 0

        for idx, row in enumerate(rows, start=2):
            try:
                row = _normalize_row(row)
                # Résoudre le PDV par numero_pdv ou pdv_id
                numero_pdv = str(row.get('numero_pdv') or row.get('numero') or '').strip()
                if not numero_pdv and row.get('pdv_id'):
                    pdv_id = _safe_int(row.get('pdv_id'))
                else:
                    # Normaliser le numero_pdv (enlever .0 si float)
                    try:
                        numero_pdv = str(int(float(numero_pdv))) if numero_pdv else ''
                    except:
                        pass
                    pdv_id = pdv_map.get(numero_pdv)

                if not pdv_id:
                    errors.append(f"Ligne {idx}: PDV '{numero_pdv}' non trouvé")
                    continue

                annee = _safe_int(row.get('annee'), 2026)
                semaine = _parse_semaine(row.get('semaine')) or _safe_int(row.get('semaine'), 1)
                ca = _safe_float(row.get('ca'))
                nb_operations = _safe_int(row.get('nb_operations'))
                nb_depots = _safe_int(row.get('nb_depots'))
                montant_depots = _safe_float(row.get('montant_depots'))
                nb_retraits = _safe_int(row.get('nb_retraits'))
                montant_retraits = _safe_float(row.get('montant_retraits'))
                est_actif = _safe_bool(row.get('est_actif', ca > 0))

                # CA précédent automatique
                prev = db.query(WeeklyPerformance).filter(
                    WeeklyPerformance.pdv_id == pdv_id,
                    WeeklyPerformance.annee == annee,
                    WeeklyPerformance.semaine == (semaine - 1),
                    WeeklyPerformance.indicateur == indicateur
                ).first()
                ca_prev = prev.ca if prev else _safe_float(row.get('ca_semaine_precedente'))
                taux_variation = ((ca - ca_prev) / ca_prev * 100) if ca_prev > 0 else 0.0

                existing = db.query(WeeklyPerformance).filter(
                    WeeklyPerformance.pdv_id == pdv_id,
                    WeeklyPerformance.annee == annee,
                    WeeklyPerformance.semaine == semaine,
                    WeeklyPerformance.indicateur == indicateur
                ).first()

                if existing:
                    existing.ca = ca
                    existing.nb_operations = nb_operations
                    existing.nb_depots = nb_depots
                    existing.montant_depots = montant_depots
                    existing.nb_retraits = nb_retraits
                    existing.montant_retraits = montant_retraits
                    existing.est_actif = est_actif
                    existing.ca_semaine_precedente = ca_prev
                    existing.taux_variation = taux_variation
                    updated_count += 1
                else:
                    db.add(WeeklyPerformance(
                        pdv_id=pdv_id,
                        annee=annee,
                        semaine=semaine,
                        ca=ca,
                        nb_operations=nb_operations,
                        nb_depots=nb_depots,
                        montant_depots=montant_depots,
                        nb_retraits=nb_retraits,
                        montant_retraits=montant_retraits,
                        est_actif=est_actif,
                        ca_semaine_precedente=ca_prev,
                        taux_variation=taux_variation,
                        indicateur=indicateur
                    ))
                    created_count += 1
                total += 1
            except Exception as e:
                errors.append(f"Ligne {idx}: {str(e)}")

        # ── PDVs absents pour chaque semaine = inactifs avec ca=0 ───────────────
        # On regroupe les PDV présents PAR semaine, puis on crée les inactifs pour chaque semaine
        if total > 0:
            # Construire un dict: (annee, semaine) -> set(pdv_ids présents)
            semaine_pdv_map = {}
            for row in rows:
                r = _normalize_row(row)
                numero_pdv = str(r.get('numero_pdv') or r.get('numero') or '').strip()
                try:
                    numero_pdv = str(int(float(numero_pdv))) if numero_pdv else ''
                except:
                    pass
                pid = pdv_map.get(numero_pdv)
                if not pid:
                    continue
                a = _safe_int(r.get('annee'))
                s = _parse_semaine(r.get('semaine')) or _safe_int(r.get('semaine'))
                if a > 0 and s > 0:
                    key = (a, s)
                    if key not in semaine_pdv_map:
                        semaine_pdv_map[key] = set()
                    semaine_pdv_map[key].add(pid)

            inactif_count = 0
            all_pdv_ids = set(pdv_map.values())
            for (annee_ref, semaine_ref), pdv_ids_presents in semaine_pdv_map.items():
                pdv_ids_absents = all_pdv_ids - pdv_ids_presents
                for pdv_id_all in pdv_ids_absents:
                    existing_inactif = db.query(WeeklyPerformance).filter(
                        WeeklyPerformance.pdv_id == pdv_id_all,
                        WeeklyPerformance.annee == annee_ref,
                        WeeklyPerformance.semaine == semaine_ref,
                        WeeklyPerformance.indicateur == indicateur
                    ).first()
                    if not existing_inactif:
                        db.add(WeeklyPerformance(
                            pdv_id=pdv_id_all,
                            annee=annee_ref,
                            semaine=semaine_ref,
                            ca=0.0,
                            nb_operations=0,
                            nb_depots=0,
                            montant_depots=0.0,
                            nb_retraits=0,
                            montant_retraits=0.0,
                            est_actif=False,
                            ca_semaine_precedente=0.0,
                            taux_variation=0.0,
                            indicateur=indicateur
                        ))
                        inactif_count += 1
                    else:
                        existing_inactif.est_actif = False
                        existing_inactif.ca = 0.0
                        existing_inactif.nb_operations = 0

        # Recalculer les variations pour toutes les semaines importées
        if total > 0:
            semaines_importees = sorted(semaine_pdv_map.keys()) if 'semaine_pdv_map' in locals() else []
            for annee_ref, semaine_ref in semaines_importees:
                perfs_semaine = db.query(WeeklyPerformance).filter(
                    WeeklyPerformance.annee == annee_ref,
                    WeeklyPerformance.semaine == semaine_ref,
                    WeeklyPerformance.indicateur == indicateur
                ).all()
                for perf in perfs_semaine:
                    prev = db.query(WeeklyPerformance).filter(
                        WeeklyPerformance.pdv_id == perf.pdv_id,
                        WeeklyPerformance.annee == annee_ref,
                        WeeklyPerformance.semaine == (semaine_ref - 1),
                        WeeklyPerformance.indicateur == indicateur
                    ).first()
                    ca_prev = prev.ca if prev else 0.0
                    perf.ca_semaine_precedente = ca_prev
                    perf.taux_variation = ((perf.ca - ca_prev) / ca_prev * 100) if ca_prev > 0 else 0.0

        db.commit()
        return {
            "status": "success",
            "created": created_count,
            "updated": updated_count,
            "inactifs_auto": inactif_count if total > 0 else 0,
            "errors": errors,
            "total_rows": total,
            "message": f"{created_count} créés, {updated_count} mis à jour, {inactif_count if total > 0 else 0} inactifs auto pour indicateur {indicateur}"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Erreur traitement fichier: {str(e)}")

@router.post("/performance/monthly")
def bulk_import_monthly(
    file: UploadFile = File(...),
    indicateur: Optional[str] = Query(default='OMY'),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Bulk import monthly performance data (Excel ou CSV, numero_pdv)"""
    try:
        contents = file.file.read()
        rows = _parse_file_to_rows(contents, file.filename)

        # Cache numero_pdv -> pdv_id
        all_pdvs = db.query(PDV.id, PDV.numero_pdv).all()
        pdv_map = {str(p.numero_pdv): p.id for p in all_pdvs}

        created_count = 0
        updated_count = 0
        errors = []
        total = 0

        for idx, row in enumerate(rows, start=2):
            try:
                row = _normalize_row(row)
                numero_pdv = str(row.get('numero_pdv') or row.get('numero') or '').strip()
                if not numero_pdv and row.get('pdv_id'):
                    pdv_id = _safe_int(row.get('pdv_id'))
                else:
                    try:
                        numero_pdv = str(int(float(numero_pdv))) if numero_pdv else ''
                    except:
                        pass
                    pdv_id = pdv_map.get(numero_pdv)

                if not pdv_id:
                    errors.append(f"Ligne {idx}: PDV '{numero_pdv}' non trouvé")
                    continue

                annee = _safe_int(row.get('annee'), 2026)
                mois = _parse_mois(row.get('mois')) or _safe_int(row.get('mois'), 1)
                ca = _safe_float(row.get('ca'))
                nb_operations = _safe_int(row.get('nb_operations'))
                nb_depots = _safe_int(row.get('nb_depots'))
                montant_depots = _safe_float(row.get('montant_depots'))
                nb_retraits = _safe_int(row.get('nb_retraits'))
                montant_retraits = _safe_float(row.get('montant_retraits'))
                est_actif = _safe_bool(row.get('est_actif', ca > 0))

                # CA précédent automatique
                prev_mois = mois - 1 if mois > 1 else 12
                prev_annee = annee if mois > 1 else annee - 1
                prev = db.query(MonthlyPerformance).filter(
                    MonthlyPerformance.pdv_id == pdv_id,
                    MonthlyPerformance.annee == prev_annee,
                    MonthlyPerformance.mois == prev_mois,
                    MonthlyPerformance.indicateur == indicateur
                ).first()
                ca_prev = prev.ca if prev else _safe_float(row.get('ca_mois_precedent'))
                taux_variation = ((ca - ca_prev) / ca_prev * 100) if ca_prev > 0 else 0.0

                existing = db.query(MonthlyPerformance).filter(
                    MonthlyPerformance.pdv_id == pdv_id,
                    MonthlyPerformance.annee == annee,
                    MonthlyPerformance.mois == mois,
                    MonthlyPerformance.indicateur == indicateur
                ).first()

                if existing:
                    existing.ca = ca
                    existing.nb_operations = nb_operations
                    existing.nb_depots = nb_depots
                    existing.montant_depots = montant_depots
                    existing.nb_retraits = nb_retraits
                    existing.montant_retraits = montant_retraits
                    existing.est_actif = est_actif
                    existing.ca_mois_precedent = ca_prev
                    existing.taux_variation = taux_variation
                    updated_count += 1
                else:
                    db.add(MonthlyPerformance(
                        pdv_id=pdv_id,
                        annee=annee,
                        mois=mois,
                        ca=ca,
                        nb_operations=nb_operations,
                        nb_depots=nb_depots,
                        montant_depots=montant_depots,
                        nb_retraits=nb_retraits,
                        montant_retraits=montant_retraits,
                        est_actif=est_actif,
                        ca_mois_precedent=ca_prev,
                        taux_variation=taux_variation,
                        indicateur=indicateur
                    ))
                    created_count += 1
                total += 1
            except Exception as e:
                errors.append(f"Ligne {idx}: {str(e)}")

        # ── PDVs absents du fichier = inactifs avec ca=0 ──────────────────────
        if total > 0:
            # Construire un dict: (annee, mois) -> set(pdv_ids présents)
            mois_pdv_map = {}
            for row in rows:
                r = _normalize_row(row)
                numero_pdv = str(r.get('numero_pdv') or r.get('numero') or '').strip()
                try:
                    numero_pdv = str(int(float(numero_pdv))) if numero_pdv else ''
                except:
                    pass
                pid = pdv_map.get(numero_pdv)
                if not pid:
                    continue
                a = _safe_int(r.get('annee'))
                m = _parse_mois(r.get('mois')) or _safe_int(r.get('mois'))
                if a > 0 and m > 0:
                    key = (a, m)
                    if key not in mois_pdv_map:
                        mois_pdv_map[key] = set()
                    mois_pdv_map[key].add(pid)

            inactif_count = 0
            all_pdv_ids = set(pdv_map.values())
            for (annee_ref, mois_ref), pdv_ids_presents in mois_pdv_map.items():
                pdv_ids_absents = all_pdv_ids - pdv_ids_presents
                prev_mois = mois_ref - 1 if mois_ref > 1 else 12
                prev_annee = annee_ref if mois_ref > 1 else annee_ref - 1
                for pdv_id_all in pdv_ids_absents:
                    existing_inactif = db.query(MonthlyPerformance).filter(
                        MonthlyPerformance.pdv_id == pdv_id_all,
                        MonthlyPerformance.annee == annee_ref,
                        MonthlyPerformance.mois == mois_ref,
                        MonthlyPerformance.indicateur == indicateur
                    ).first()
                    prev = db.query(MonthlyPerformance).filter(
                        MonthlyPerformance.pdv_id == pdv_id_all,
                        MonthlyPerformance.annee == prev_annee,
                        MonthlyPerformance.mois == prev_mois,
                        MonthlyPerformance.indicateur == indicateur
                    ).first()
                    ca_prev_inactif = prev.ca if prev else 0.0
                    taux_var_inactif = -100.0 if ca_prev_inactif > 0 else 0.0

                    if not existing_inactif:
                        db.add(MonthlyPerformance(
                            pdv_id=pdv_id_all,
                            annee=annee_ref,
                            mois=mois_ref,
                            ca=0.0,
                            nb_operations=0,
                            nb_depots=0,
                            montant_depots=0.0,
                            nb_retraits=0,
                            montant_retraits=0.0,
                            est_actif=False,
                            ca_mois_precedent=ca_prev_inactif,
                            taux_variation=taux_var_inactif,
                            indicateur=indicateur
                        ))
                        inactif_count += 1
                    else:
                        existing_inactif.est_actif = False
                        existing_inactif.ca = 0.0
                        existing_inactif.nb_operations = 0
                        existing_inactif.ca_mois_precedent = ca_prev_inactif
                        existing_inactif.taux_variation = taux_var_inactif

        db.commit()
        return {
            "status": "success",
            "created": created_count,
            "updated": updated_count,
            "inactifs_auto": inactif_count if total > 0 and annee_ref > 0 else 0,
            "errors": errors,
            "total_rows": total,
            "message": f"{created_count} créés, {updated_count} mis à jour, {inactif_count if total > 0 and annee_ref > 0 else 0} inactifs auto pour indicateur {indicateur}"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Erreur traitement fichier: {str(e)}")

@router.get("/performance/pdv/{pdv_id}/weekly", response_model=List[WeeklyPerformanceResponse])
def get_pdv_weekly_history(
    pdv_id: int,
    db: Session = Depends(get_db),
    annee: int = None,
    limit: int = 52
):
    """Get weekly history for a PDV"""
    pdv = db.query(PDV).filter(PDV.id == pdv_id).first()
    if not pdv:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="PDV not found"
        )
    
    query = db.query(WeeklyPerformance).filter(
        WeeklyPerformance.pdv_id == pdv_id
    )
    
    if annee:
        query = query.filter(WeeklyPerformance.annee == annee)
    
    performances = query.order_by(
        WeeklyPerformance.annee.desc(),
        WeeklyPerformance.semaine.desc()
    ).limit(limit).all()
    
    return [WeeklyPerformanceResponse.from_orm(p) for p in performances]

@router.get("/performance/pdv/{pdv_id}/monthly", response_model=List[MonthlyPerformanceResponse])
def get_pdv_monthly_history(
    pdv_id: int,
    db: Session = Depends(get_db),
    annee: int = None,
    limit: int = 24
):
    """Get monthly history for a PDV"""
    pdv = db.query(PDV).filter(PDV.id == pdv_id).first()
    if not pdv:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="PDV not found"
        )
    
    query = db.query(MonthlyPerformance).filter(
        MonthlyPerformance.pdv_id == pdv_id
    )
    
    if annee:
        query = query.filter(MonthlyPerformance.annee == annee)
    
    performances = query.order_by(
        MonthlyPerformance.annee.desc(),
        MonthlyPerformance.mois.desc()
    ).limit(limit).all()
    
    return [MonthlyPerformanceResponse.from_orm(p) for p in performances]


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY ENDPOINT (utilisé par Kaabu, Nafama, OMY dashboards)
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/performance/monthly/summary")
def monthly_summary(
    annee: int = Query(...),
    mois: int = Query(...),
    indicateur: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """
    Résumé mensuel des performances, filtré par indicateur si fourni.
    Retourne: ca_total, montant_ca, commission_pdg, commission_revendeur,
              pdvs_actifs, pdvs_inactifs, ca_moyen, by_zone, by_type, evolution
    """
    q = db.query(MonthlyPerformance, PDV).join(PDV, MonthlyPerformance.pdv_id == PDV.id).filter(
        MonthlyPerformance.annee == annee,
        MonthlyPerformance.mois == mois,
        PDV.statut != 'DESACTIVE',
    )
    if indicateur:
        q = q.filter(MonthlyPerformance.indicateur == indicateur)

    pairs = q.all()

    def _mt(p):
        mt = getattr(p, 'montant_transaction', None)
        if mt and mt > 0: return mt
        return (p.montant_depots or 0.0) + (p.montant_retraits or 0.0)

    n = len(pairs)
    ca_total      = sum(_mt(p) for p, _ in pairs)
    montant_ca    = sum(getattr(p, 'montant_ca', None) or 0.0 for p, _ in pairs)
    comm_pdg      = sum(getattr(p, 'commission_pdg', None) or 0.0 for p, _ in pairs)
    comm_rev      = sum(getattr(p, 'commission_revendeur', None) or 0.0 for p, _ in pairs)
    pdvs_actifs   = sum(1 for p, _ in pairs if p.est_actif)
    pdvs_inactifs = n - pdvs_actifs
    ca_moyen      = ca_total / n if n > 0 else 0
    ratio_ca      = round((montant_ca / ca_total * 100) if ca_total > 0 else 0, 2)

    # Par zone
    by_zone: dict = {}
    for p, pdv in pairs:
        z = pdv.zone or "Inconnue"
        if z not in by_zone:
            by_zone[z] = {"zone": z, "ca_total": 0, "pdvs": 0, "actifs": 0}
        by_zone[z]["ca_total"] += _mt(p)
        by_zone[z]["pdvs"] += 1
        if p.est_actif: by_zone[z]["actifs"] += 1
    by_zone_list = sorted(by_zone.values(), key=lambda x: x["ca_total"], reverse=True)

    # Par type PDV
    by_type: dict = {}
    for p, pdv in pairs:
        t = pdv.type_pdv.value
        if t not in by_type:
            by_type[t] = {"type": t, "ca_total": 0, "pdvs": 0}
        by_type[t]["ca_total"] += _mt(p)
        by_type[t]["pdvs"] += 1

    # Evolution 12 derniers mois (même indicateur)
    from sqlalchemy import and_
    evo_q = db.query(
        MonthlyPerformance.annee,
        MonthlyPerformance.mois,
        db.query(MonthlyPerformance.montant_transaction).label('mt'),
    )
    # Simplification : agrégation Python sur les 12 mois
    evolution: list = []
    MOIS_NOM = ['','Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
    for m in range(1, 13):
        a = annee if m <= mois else annee - 1
        mo = m
        evo_pairs_q = db.query(MonthlyPerformance).filter(
            MonthlyPerformance.annee == a,
            MonthlyPerformance.mois == mo,
        )
        if indicateur:
            evo_pairs_q = evo_pairs_q.filter(MonthlyPerformance.indicateur == indicateur)
        evo_perfs = evo_pairs_q.all()
        total_m = sum(_mt(p) for p in evo_perfs)
        if total_m > 0 or (a == annee and mo <= mois):
            evolution.append({
                "mois": mo, "annee": a,
                "label": f"{MOIS_NOM[mo]} {a}",
                "ca_total": round(total_m, 0),
                "montant_ca": round(sum(getattr(p,'montant_ca',None) or 0 for p in evo_perfs), 0),
                "commission_pdg": round(sum(getattr(p,'commission_pdg',None) or 0 for p in evo_perfs), 2),
                "pdvs_actifs": sum(1 for p in evo_perfs if p.est_actif),
            })

    # Top PDV
    top_pdv = sorted(
        [{"pdv_id": pdv.id, "numero_pdv": pdv.numero_pdv, "nom": pdv.nom,
          "zone": pdv.zone, "superviseur": pdv.superviseur,
          "ca_total": round(_mt(p), 0),
          "montant_ca": round(getattr(p,'montant_ca',None) or 0, 0),
          "commission_pdg": round(getattr(p,'commission_pdg',None) or 0, 2),
          "commission_revendeur": round(getattr(p,'commission_revendeur',None) or 0, 2),
          "nb_operations": p.nb_operations, "est_actif": p.est_actif,
          "taux_variation": p.taux_variation}
         for p, pdv in pairs if p.est_actif],
        key=lambda x: x["ca_total"], reverse=True
    )[:20]

    return {
        "annee": annee, "mois": mois, "indicateur": indicateur,
        "n_pdv": n,
        # KPIs principaux
        "ca_total":    round(ca_total, 0),          # Montant Transaction (compatibilité)
        "montant_transaction": round(ca_total, 0),  # Nouveau nom
        "montant_ca":  round(montant_ca, 0),         # CA Orange
        "commission_pdg": round(comm_pdg, 2),
        "commission_revendeur": round(comm_rev, 2),
        "ratio_ca_transaction": ratio_ca,
        "ca_moyen":    round(ca_moyen, 0),
        "pdvs_actifs": pdvs_actifs,
        "pdvs_inactifs": pdvs_inactifs,
        "taux_activite": round(pdvs_actifs / n * 100 if n > 0 else 0, 1),
        # Ventilations
        "by_zone": by_zone_list,
        "by_type": list(by_type.values()),
        "evolution": evolution,
        "top_pdv": top_pdv,
    }


@router.get("/performance/weekly/summary")
def weekly_summary(
    annee: int = Query(...),
    semaine: int = Query(...),
    indicateur: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """Résumé hebdomadaire des performances."""
    q = db.query(WeeklyPerformance, PDV).join(PDV, WeeklyPerformance.pdv_id == PDV.id).filter(
        WeeklyPerformance.annee == annee,
        WeeklyPerformance.semaine == semaine,
        PDV.statut != 'DESACTIVE',
    )
    if indicateur:
        q = q.filter(WeeklyPerformance.indicateur == indicateur)

    pairs = q.all()

    def _wt(p):
        mt = getattr(p, 'montant_transaction', None)
        if mt and mt > 0: return mt
        return (p.montant_depots or 0.0) + (p.montant_retraits or 0.0)

    n = len(pairs)
    ca_total   = sum(_wt(p) for p, _ in pairs)
    montant_ca = sum(getattr(p, 'montant_ca', None) or 0.0 for p, _ in pairs)
    comm_pdg   = sum(getattr(p, 'commission_pdg', None) or 0.0 for p, _ in pairs)
    comm_rev   = sum(getattr(p, 'commission_revendeur', None) or 0.0 for p, _ in pairs)
    pdvs_actifs = sum(1 for p, _ in pairs if p.est_actif)

    by_zone: dict = {}
    for p, pdv in pairs:
        z = pdv.zone or "Inconnue"
        if z not in by_zone:
            by_zone[z] = {"zone": z, "ca_total": 0, "pdvs": 0}
        by_zone[z]["ca_total"] += _wt(p)
        by_zone[z]["pdvs"] += 1

    top_pdv = sorted(
        [{"pdv_id": pdv.id, "numero_pdv": pdv.numero_pdv, "nom": pdv.nom,
          "zone": pdv.zone, "ca_total": round(_wt(p), 0),
          "montant_ca": round(getattr(p,'montant_ca',None) or 0, 0),
          "commission_pdg": round(getattr(p,'commission_pdg',None) or 0, 2),
          "commission_revendeur": round(getattr(p,'commission_revendeur',None) or 0, 2),
          "nb_operations": p.nb_operations, "est_actif": p.est_actif}
         for p, pdv in pairs if p.est_actif],
        key=lambda x: x["ca_total"], reverse=True
    )[:20]

    return {
        "annee": annee, "semaine": semaine, "indicateur": indicateur,
        "n_pdv": n,
        "ca_total": round(ca_total, 0),
        "montant_transaction": round(ca_total, 0),
        "montant_ca": round(montant_ca, 0),
        "commission_pdg": round(comm_pdg, 2),
        "commission_revendeur": round(comm_rev, 2),
        "ca_moyen": round(ca_total / n if n > 0 else 0, 0),
        "pdvs_actifs": pdvs_actifs,
        "pdvs_inactifs": n - pdvs_actifs,
        "taux_activite": round(pdvs_actifs / n * 100 if n > 0 else 0, 1),
        "by_zone": sorted(by_zone.values(), key=lambda x: x["ca_total"], reverse=True),
        "top_pdv": top_pdv,
    }


# ─────────────────────────────────────────────────────────────────────────────
# IMPORT DEPUIS FICHIER EXPORT ORANGE
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/performance/import-export-orange")
async def import_export_orange(
    file: UploadFile = File(...),
    mode: str = Form("mensuel"),          # mensuel | hebdo
    annee: Optional[int] = Form(None),
    mois: Optional[int] = Form(None),
    semaine: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(require_admin),
):
    """
    Import depuis le fichier EXPORT Orange (EXPORT_XXXXXXXXX.xlsx).

    Structure attendue :
      Col A: Numero revendeur | Col B: Grade | Col F: Service
      Col G: Nombre transaction | Col H: Montant transaction
      Col I: Transaction CA | Col J: Commission PDG | Col K: Commission revendeur
      Col N: Date transaction (DD/MM/YYYY)

    mode='mensuel' → agrège par (PDV, mois)
    mode='hebdo'   → agrège par (PDV, semaine ISO)
    """
    import openpyxl
    from io import BytesIO
    from datetime import datetime
    from collections import defaultdict

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Fichier Excel (.xlsx/.xls) requis")

    content = await file.read()

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Impossible de lire le fichier : {e}")

    # Lecture entêtes
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell or '').strip().lower())

    # Mapping colonnes
    aliases = {
        'numero revendeur': 'numero', 'grade': 'grade', 'service': 'service',
        'nombre transction': 'nb_trans', 'nombre transaction': 'nb_trans',
        'montant transaction': 'mt_trans', 'transaction ca': 'trans_ca',
        'commission pdg': 'comm_pdg', 'commission revendeur': 'comm_rev',
        'date transaction': 'date',
    }
    col = {}
    for i, h in enumerate(headers):
        key = aliases.get(h)
        if key:
            col[key] = i

    required = ['numero', 'service', 'nb_trans', 'mt_trans', 'date']
    missing = [r for r in required if r not in col]
    if missing:
        raise HTTPException(400, f"Colonnes manquantes : {missing}. Trouvées : {headers}")

    # Helpers
    def _f(val):
        try: return float(str(val or 0).replace(' ', '').replace(',', '.'))
        except: return 0.0
    def _i(val):
        try: return int(float(str(val or 0)))
        except: return 0
    def _date(val):
        s = str(val or '').strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try: return datetime.strptime(s, fmt)
            except: continue
        return None

    # Chargement PDV
    pdv_map = {str(p.numero_pdv).strip(): p.id for p in db.query(PDV).all()}

    # Agrégation
    data = defaultdict(lambda: {
        'nb_depots': 0, 'mt_depots': 0.0, 'nb_retraits': 0, 'mt_retraits': 0.0,
        'trans_ca': 0.0, 'comm_pdg': 0.0, 'comm_rev': 0.0,
    })

    skipped_date = skipped_pdv = total = 0
    not_found = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col['numero']] is None:
            continue
        total += 1
        numero = str(row[col['numero']]).strip()
        d = _date(row[col['date']])
        if d is None:
            skipped_date += 1
            continue

        annee_row = d.year
        if annee and annee_row != annee:
            continue
        mois_row = d.month
        if mois and mois_row != mois:
            continue
        sem_row = d.isocalendar()[1]
        if semaine and sem_row != semaine:
            continue

        key = (numero, annee_row, mois_row) if mode == 'mensuel' else (numero, annee_row, sem_row)
        svc = str(row[col['service']] or '').strip().upper()
        nb = _i(row[col['nb_trans']])
        mt = _f(row[col['mt_trans']])
        ca = _f(row[col['trans_ca']]) if 'trans_ca' in col else 0.0
        cpdg = _f(row[col['comm_pdg']]) if 'comm_pdg' in col else 0.0
        crev = _f(row[col['comm_rev']]) if 'comm_rev' in col else 0.0

        rec = data[key]
        rec['trans_ca'] += ca
        rec['comm_pdg'] += cpdg
        rec['comm_rev'] += crev
        if svc == 'CASHIN':
            rec['nb_depots'] += nb; rec['mt_depots'] += mt
        elif svc == 'CASHOUT':
            rec['nb_retraits'] += nb; rec['mt_retraits'] += mt

    # Upsert en base
    created = updated = 0
    for (numero, annee_v, periode_v), rec in data.items():
        pdv_id = pdv_map.get(numero)
        if pdv_id is None:
            not_found.add(numero); skipped_pdv += 1
            continue

        mt_total = rec['mt_depots'] + rec['mt_retraits']
        nb_ops   = rec['nb_depots'] + rec['nb_retraits']
        ratio    = round((rec['trans_ca'] / mt_total * 100) if mt_total > 0 else 0.0, 4)

        if mode == 'mensuel':
            existing = db.query(MonthlyPerformance).filter_by(
                pdv_id=pdv_id, annee=annee_v, mois=periode_v
            ).first()
            if existing:
                existing.nb_depots=rec['nb_depots']; existing.montant_depots=rec['mt_depots']
                existing.nb_retraits=rec['nb_retraits']; existing.montant_retraits=rec['mt_retraits']
                existing.nb_operations=nb_ops; existing.montant_transaction=mt_total
                existing.montant_ca=rec['trans_ca']; existing.commission_pdg=rec['comm_pdg']
                existing.commission_revendeur=rec['comm_rev']; existing.ratio_ca_transaction=ratio
                existing.est_actif=nb_ops > 0; existing.ca=mt_total
                updated += 1
            else:
                db.add(MonthlyPerformance(
                    pdv_id=pdv_id, annee=annee_v, mois=periode_v,
                    nb_depots=rec['nb_depots'], montant_depots=rec['mt_depots'],
                    nb_retraits=rec['nb_retraits'], montant_retraits=rec['mt_retraits'],
                    nb_operations=nb_ops, montant_transaction=mt_total,
                    montant_ca=rec['trans_ca'], commission_pdg=rec['comm_pdg'],
                    commission_revendeur=rec['comm_rev'], ratio_ca_transaction=ratio,
                    est_actif=nb_ops > 0, ca=mt_total,
                ))
                created += 1
        else:
            existing = db.query(WeeklyPerformance).filter_by(
                pdv_id=pdv_id, annee=annee_v, semaine=periode_v
            ).first()
            if existing:
                existing.nb_depots=rec['nb_depots']; existing.montant_depots=rec['mt_depots']
                existing.nb_retraits=rec['nb_retraits']; existing.montant_retraits=rec['mt_retraits']
                existing.nb_operations=nb_ops; existing.montant_transaction=mt_total
                existing.montant_ca=rec['trans_ca']; existing.commission_pdg=rec['comm_pdg']
                existing.commission_revendeur=rec['comm_rev']; existing.ratio_ca_transaction=ratio
                existing.est_actif=nb_ops > 0; existing.ca=mt_total
                updated += 1
            else:
                db.add(WeeklyPerformance(
                    pdv_id=pdv_id, annee=annee_v, semaine=periode_v,
                    nb_depots=rec['nb_depots'], montant_depots=rec['mt_depots'],
                    nb_retraits=rec['nb_retraits'], montant_retraits=rec['mt_retraits'],
                    nb_operations=nb_ops, montant_transaction=mt_total,
                    montant_ca=rec['trans_ca'], commission_pdg=rec['comm_pdg'],
                    commission_revendeur=rec['comm_rev'], ratio_ca_transaction=ratio,
                    est_actif=nb_ops > 0, ca=mt_total,
                ))
                created += 1

    db.commit()

    # ── Créer des entrées est_actif=False pour les PDVs absents ──────────────
    # Collecter toutes les périodes importées
    inactif_created = 0
    if mode == 'mensuel':
        # Pour chaque (annee, mois) importé, créer des entrées inactives pour les PDVs absents
        periodes = set((annee_v, periode_v) for (_, annee_v, periode_v) in data.keys())
        for (annee_v, mois_v) in periodes:
            pdv_ids_actifs = {
                pdv_map[num] for (num, a, m) in data.keys()
                if a == annee_v and m == mois_v and num in pdv_map
            }
            all_pdv_ids = set(pdv_map.values())
            pdv_ids_absents = all_pdv_ids - pdv_ids_actifs
            for pid in pdv_ids_absents:
                existing = db.query(MonthlyPerformance).filter_by(
                    pdv_id=pid, annee=annee_v, mois=mois_v
                ).first()
                if not existing:
                    db.add(MonthlyPerformance(
                        pdv_id=pid, annee=annee_v, mois=mois_v,
                        nb_depots=0, montant_depots=0.0,
                        nb_retraits=0, montant_retraits=0.0,
                        nb_operations=0, montant_transaction=0.0,
                        montant_ca=0.0, commission_pdg=0.0,
                        commission_revendeur=0.0, ratio_ca_transaction=0.0,
                        est_actif=False, ca=0.0,
                    ))
                    inactif_created += 1
    else:
        # Hebdo: pour chaque (annee, semaine) importée
        periodes = set((annee_v, periode_v) for (_, annee_v, periode_v) in data.keys())
        for (annee_v, sem_v) in periodes:
            pdv_ids_actifs = {
                pdv_map[num] for (num, a, s) in data.keys()
                if a == annee_v and s == sem_v and num in pdv_map
            }
            all_pdv_ids = set(pdv_map.values())
            pdv_ids_absents = all_pdv_ids - pdv_ids_actifs
            for pid in pdv_ids_absents:
                existing = db.query(WeeklyPerformance).filter_by(
                    pdv_id=pid, annee=annee_v, semaine=sem_v
                ).first()
                if not existing:
                    db.add(WeeklyPerformance(
                        pdv_id=pid, annee=annee_v, semaine=sem_v,
                        nb_depots=0, montant_depots=0.0,
                        nb_retraits=0, montant_retraits=0.0,
                        nb_operations=0, montant_transaction=0.0,
                        montant_ca=0.0, commission_pdg=0.0,
                        commission_revendeur=0.0, ratio_ca_transaction=0.0,
                        est_actif=False, ca=0.0,
                    ))
                    inactif_created += 1

    db.commit()
    # ─────────────────────────────────────────────────────────────────────────

    return {
        "success": True,
        "mode": mode,
        "total_lignes": total,
        "pdv_periodes_traitees": len(data),
        "created": created,
        "updated": updated,
        "inactifs_crees": inactif_created,
        "skipped_pdv": skipped_pdv,
        "skipped_date": skipped_date,
        "pdv_non_trouves": len(not_found),
        "exemples_non_trouves": list(not_found)[:10],
    }
